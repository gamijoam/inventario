from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, desc
from typing import Optional
from datetime import datetime, date
from decimal import Decimal
from ..database.db import get_db
from ..models import models
from ..dependencies import admin_only

router = APIRouter(
    prefix="/reports",
    tags=["reports"],
    dependencies=[Depends(admin_only)]  # 🔒 ADMIN ONLY - Financial data is sensitive
)

@router.get("/dashboard/financials")
def get_dashboard_financials(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """
    Financial metrics for dashboard - real money collected by currency
    
    Returns:
    - sales_by_currency: List of totals grouped by currency (USD, COP, VES, etc.)
    - total_sales_base_usd: Sum of all sales converted to USD base
    - profit_estimated: Estimated profit (sales - costs)
    """
    # Default to today if no dates provided
    if not start_date:
        start_date = date.today()
    if not end_date:
        end_date = date.today()
    
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # Query SalePayment grouped by currency
    # Note: We include ALL sales, even if they have returns. Returns are subtracted separately.
    query = db.query(
        models.SalePayment.currency,
        func.sum(models.SalePayment.amount).label('total_collected'),
        func.count(models.SalePayment.id).label('payment_count')
    ).join(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    )
    
    # Group by currency
    results = query.group_by(models.SalePayment.currency).all()
    
    # NEW: Query Returns (CashMovements of type "RETURN")
    # This captures the actual money leaving the drawer for refunds
    returns_query = db.query(
        models.CashMovement.currency,
        func.sum(models.CashMovement.amount).label('total_refunded')
    ).filter(
        models.CashMovement.date >= start_dt,
        models.CashMovement.date <= end_dt,
        models.CashMovement.type == "RETURN"
    ).group_by(models.CashMovement.currency).all()
    
    # Convert returns to dict for easy lookup
    returns_map = {r[0] or "USD": r[1] for r in returns_query}
    
    # Format sales by currency
    sales_by_currency = []
    total_sales_base_usd = Decimal("0.00")
    
    for currency, total_collected, count in results:
        # Safety check: total_collected usually isn't None but SQL sum can be
        if total_collected is None:
            total_collected = Decimal("0.00")
        else:
            total_collected = Decimal(str(total_collected))
            
        # Subtract returns for this currency
        refunds = returns_map.get(currency or "USD", Decimal("0.00"))
        # Ensure refunds is Decimal (just in case)
        if refunds is None: 
            refunds = Decimal("0.00")
        else:
            refunds = Decimal(str(refunds))
            
        net_collected = total_collected - refunds
        
        sales_by_currency.append({
            "currency": currency or "USD",
            "total_collected": float(round(net_collected, 2)),
            "count": count,
            "returns": float(round(refunds, 2)) # Optional: show returns
        })
        
        # Convert to USD for base total
        # If currency is USD, add directly; otherwise use exchange rate
        if currency == "USD":
            total_sales_base_usd += Decimal(str(net_collected))
        else:
            # Get average exchange rate for the period
            avg_rate = db.query(func.avg(models.SalePayment.exchange_rate)).filter(
                models.SalePayment.currency == currency
            ).join(models.Sale).filter(
                models.Sale.date >= start_dt,
                models.Sale.date <= end_dt
            ).scalar()
            
            # Safety checks for avg_rate
            if avg_rate is None or avg_rate == 0:
                avg_rate = Decimal("1.0")
            
            # Convert to USD
            total_sales_base_usd += Decimal(str(net_collected)) / Decimal(str(avg_rate))
    
    # Calculate profit estimation (Sales - Costs)
    # Get all sale details for the period
    sale_details = db.query(models.SaleDetail).join(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).all()
    
    total_cost = Decimal("0.00")
    total_revenue = Decimal("0.00")
    
    for detail in sale_details:
        if detail.subtotal is not None:
            total_revenue += detail.subtotal
        
        # HISTORICAL COST LOGIC
        # Use cost_at_sale if available (new sales), otherwise fallback to current product cost (legacy)
        cost_price = detail.product.cost_price
        if detail.cost_at_sale is not None and detail.cost_at_sale > 0:
            cost_price = detail.cost_at_sale
            
        if cost_price and detail.quantity is not None:
            total_cost += Decimal(str(cost_price)) * Decimal(str(detail.quantity))
            
    # Subtract returns from revenue and add back cost (roughly)
    # Ideally we'd look at ReturnDetails for exact cost reversal, but for estimation:
    # We subtract the cash refund from revenue.
    # Cost reversal is complex because "DAMAGED" items are lost (cost remains), "GOOD" are restored (cost removed).
    # For now, let's keep it simple: Profit = (Revenue - Returns) - (Cost of Sales - Cost of Good Returns)
    # Getting exact cost of returned items:
    
    return_details = db.query(models.ReturnDetail).join(models.Return).filter(
        models.Return.date >= start_dt,
        models.Return.date <= end_dt
    ).options(joinedload(models.ReturnDetail.product)).all()
    
    total_refunds_revenue = Decimal("0.00")
    total_refunds_cost = Decimal("0.00")
    
    for rd in return_details:
        quantity = Decimal(str(rd.quantity)) if rd.quantity is not None else Decimal("0")
        unit_price = Decimal(str(rd.unit_price)) if rd.unit_price is not None else Decimal("0")
        
        total_refunds_revenue += (quantity * unit_price)
        
        # HISTORICAL RETURN COST LOGIC
        cost_price = rd.product.cost_price
        if rd.unit_cost is not None and rd.unit_cost > 0:
            cost_price = rd.unit_cost
            
        if cost_price:
            cost_price_dec = Decimal(str(cost_price))
            total_refunds_cost += (quantity * cost_price_dec)

    # Adjusted Profit
    # Revenue = (Sales Revenue - Refunds)
    # Cost = (Sales Cost - Returned Items Cost) -> Assuming returned items go back to stock (value recovered)
    # Note: If damaged, we effectively lost the cost, so we SHOULD NOT subtract it from cost (cost remains).
    # Since we don't easily know here, let's assume worst case (lost) or best case (recovered).
    # Given most returns are "didn't want it", let's assume recovered.
    
    final_revenue = total_revenue - total_refunds_revenue
    final_cost = total_cost - total_refunds_cost
    
    profit_estimated = final_revenue - final_cost
    
    return {
        "sales_by_currency": sales_by_currency,
        "total_sales_base_usd": float(round(total_sales_base_usd, 2)),
        "profit_estimated": float(round(profit_estimated, 2))
    }

@router.get("/dashboard/cashflow")
def get_dashboard_cashflow(db: Session = Depends(get_db)):
    """
    Calculate physical cash balance by currency in open cash sessions
    
    Returns real money that should be in the cash drawer:
    - Initial cash from open sessions
    - + Sales income (from SalePayment)
    - + Deposits
    - - Expenses
    - - Withdrawals  
    - - Returns/Refunds
    """
    # Get all active currencies from config
    active_currencies = db.query(models.Currency).filter(models.Currency.is_active == True).all()
    currency_codes = [c.symbol for c in active_currencies] if active_currencies else ['USD', 'Bs']
    
    # Get open cash sessions
    open_sessions = db.query(models.CashSession).filter(models.CashSession.status == "OPEN").all()
    
    if not open_sessions:
        # No open sessions, return zeros
        return {
            "balances": [{"currency": code, "initial": 0, "sales": 0, "expenses": 0, "net_balance": 0} for code in currency_codes],
            "alerts": ["No hay sesiones de caja abiertas"]
        }
    
    session_ids = [s.id for s in open_sessions]
    balances = {}
    alerts = []
    
    # Initialize balances for each currency
    for currency in currency_codes:
        balances[currency] = {
            "currency": currency,
            "initial": 0.0,
            "sales": 0.0,
            "expenses": 0.0,
            "net_balance": 0.0
        }
    
    # 1. Get initial cash from open sessions
    for session in open_sessions:
        # Check if session has multi-currency support
        if session.currencies:
            for curr in session.currencies:
                if curr.currency_symbol in balances:
                    balances[curr.currency_symbol]["initial"] += curr.initial_amount
        else:
            # Fallback to old dual-currency model
            balances["USD"]["initial"] += session.initial_cash or 0
            if "Bs" in balances:
                balances["Bs"]["initial"] += session.initial_cash_bs or 0
    
    # 2. Get sales income from SalePayment (exclude voided sales)
    voided_sale_ids = db.query(models.Return.sale_id).distinct().all()
    voided_sale_ids = [sid[0] for sid in voided_sale_ids]
    
    sales_query = db.query(
        models.SalePayment.currency,
        func.sum(models.SalePayment.amount).label('total')
    ).join(models.Sale).filter(
        models.Sale.date >= open_sessions[0].start_time  # Since first session opened
    )
    
    if voided_sale_ids:
        sales_query = sales_query.filter(models.Sale.id.notin_(voided_sale_ids))
    
    sales_by_currency = sales_query.group_by(models.SalePayment.currency).all()
    
    for currency, total in sales_by_currency:
        if currency in balances:
            balances[currency]["sales"] += total
    
    # 3. Get cash movements (expenses, deposits, withdrawals, returns)
    movements = db.query(models.CashMovement).filter(
        models.CashMovement.session_id.in_(session_ids)
    ).all()
    
    for movement in movements:
        currency = movement.currency or "USD"
        if currency not in balances:
            continue
            
        if movement.type == "DEPOSIT":
            # Deposits are income
            balances[currency]["sales"] += movement.amount
        elif movement.type in ["EXPENSE", "WITHDRAWAL"]:
            # Expenses and withdrawals reduce cash
            balances[currency]["expenses"] -= movement.amount
        elif movement.type == "RETURN":
            # Returns are refunds (reduce cash)
            balances[currency]["expenses"] -= movement.amount
    
    # 4. Calculate net balance and check for alerts
    for currency_code, data in balances.items():
        data["net_balance"] = data["initial"] + data["sales"] + data["expenses"]
        
        # Alert if negative balance
        if data["net_balance"] < 0:
            alerts.append(f"⚠️ Caja en {currency_code} tiene saldo negativo: {data['net_balance']:.2f} (Revisar)")
        
        # Round values
        data["initial"] = round(data["initial"], 2)
        data["sales"] = round(data["sales"], 2)
        data["expenses"] = round(data["expenses"], 2)
        data["net_balance"] = round(data["net_balance"], 2)
    
    # Convert to list and filter out currencies with no activity
    balance_list = [data for data in balances.values() if data["initial"] != 0 or data["sales"] != 0 or data["expenses"] != 0]
    
    # If no activity, show at least USD
    if not balance_list:
        balance_list = [balances["USD"]]
    
    return {
        "balances": balance_list,
        "alerts": alerts if alerts else []
    }

@router.get("/sales/detailed")
def get_detailed_sales_report(
    start_date: date,
    end_date: date,
    customer_id: Optional[int] = None,
    product_id: Optional[int] = None,
    payment_method: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Detailed sales report with filters"""
    # Convert dates to datetime
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    query = db.query(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    )
    
    if customer_id:
        query = query.filter(models.Sale.customer_id == customer_id)
    
    if payment_method:
        query = query.filter(models.Sale.payment_method == payment_method)
    
    sales = query.order_by(models.Sale.date.desc()).all()
    
    # Filter by product if specified
    if product_id:
        sales = [s for s in sales if any(d.product_id == product_id for d in s.details)]
    
    return sales

@router.get("/sales/summary")
def get_sales_summary(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db)
):
    """Summary statistics for sales period"""
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    sales = db.query(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).all()
    
    total_revenue = Decimal(0)
    total_revenue_bs = Decimal(0)
    
    for s in sales:
        # Robust Amount Casting
        amount = s.total_amount
        if amount is None:
            amount = Decimal(0)
        elif not isinstance(amount, Decimal):
            amount = Decimal(str(amount))
            
        total_revenue += amount
        
        # Robust BS Calculation
        if s.total_amount_bs is not None:
            bs_amount = s.total_amount_bs
            if not isinstance(bs_amount, Decimal):
                bs_amount = Decimal(str(bs_amount))
            total_revenue_bs += bs_amount
        else:
            rate = s.exchange_rate_used
            if rate is None:
                rate = Decimal("1.0")
            elif not isinstance(rate, Decimal):
                rate = Decimal(str(rate))
            
            total_revenue_bs += (amount * rate)
            
    total_transactions = len(sales)
    
    # Count by payment method
    cash_sales = Decimal(0)
    credit_sales = Decimal(0)
    
    for s in sales:
        amount = s.total_amount
        if amount is None:
            amount = Decimal(0)
        elif not isinstance(amount, Decimal):
            amount = Decimal(str(amount))
            
        if s.payment_method == "Efectivo":
            cash_sales += amount
        elif s.payment_method == "Credito":
            credit_sales += amount
    
    # Total items sold
    total_items = db.query(func.sum(models.SaleDetail.quantity)).join(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).scalar() or 0
    
    # Subtract returns
    returns = db.query(models.Return).filter(
        models.Return.date >= start_dt,
        models.Return.date <= end_dt
    ).all()
    
    total_refunded = sum((r.total_refunded or 0) for r in returns)
    total_refunded_bs = Decimal(0)
    for ret in returns:
        sale = db.query(models.Sale).filter(models.Sale.id == ret.sale_id).first()
        if sale:
            refund_amount = ret.total_refunded or 0
            exchange_rate = sale.exchange_rate_used or Decimal("1.0")
            
            # Ensure Decimal
            if not isinstance(refund_amount, Decimal):
                 refund_amount = Decimal(str(refund_amount))
            if not isinstance(exchange_rate, Decimal):
                 exchange_rate = Decimal(str(exchange_rate))
                
            refund_bs = refund_amount * exchange_rate
            total_refunded_bs += refund_bs
    
    # Adjust totals
    total_revenue -= total_refunded
    total_revenue_bs -= total_refunded_bs
    
    avg_ticket = total_revenue / total_transactions if total_transactions > 0 else Decimal(0)
    
    return {
        "total_revenue": float(total_revenue),
        "total_revenue_bs": float(total_revenue_bs),
        "total_transactions": total_transactions,
        "cash_sales": float(cash_sales),
        "credit_sales": float(credit_sales),
        "total_items_sold": float(total_items),
        "total_refunded": float(total_refunded),
        "average_ticket": float(avg_ticket)
    }

@router.get("/cash-flow")
def get_cash_flow_report(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db)
):
    """All cash movements in period"""
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # Get all cash sessions in period
    sessions = db.query(models.CashSession).filter(
        models.CashSession.start_time >= start_dt,
        models.CashSession.start_time <= end_dt
    ).all()
    
    movements = []
    for session in sessions:
        for mov in session.movements:
            movements.append({
                "date": mov.date.isoformat(),
                "session_id": session.id,
                "type": mov.type,
                "amount": mov.amount,
                "currency": mov.currency or "USD",
                "description": mov.description
            })
    
    # Also get sales
    sales = db.query(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt,
        models.Sale.is_credit == False
    ).all()
    
    for sale in sales:
        movements.append({
            "date": sale.date.isoformat(),
            "session_id": None,
            "type": "SALE",
            "amount": sale.total_amount,
            "currency": "USD",
            "description": f"Venta #{sale.id}"
        })
    
    # Sort by date
    movements.sort(key=lambda x: x["date"], reverse=True)
    
    return movements

@router.get("/top-products")
def get_top_products(
    start_date: date,
    end_date: date,
    limit: int = 10,
    by: str = "quantity",
    db: Session = Depends(get_db)
):
    """Top products by NET quantity or revenue (Gross - Returns)"""
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # 1. Get Gross Sales
    sales_query = db.query(
        models.Product.id,
        models.Product.name,
        func.sum(models.SaleDetail.quantity).label('gross_qty'),
        func.sum(models.SaleDetail.subtotal).label('gross_rev')
    ).join(models.SaleDetail).join(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).group_by(models.Product.id, models.Product.name).all()
    
    sales_map = {r[0]: {"name": r[1], "gross_qty": float(r[2]), "gross_rev": float(r[3])} for r in sales_query}
    
    # 2. Get Returns in period
    # Note: We filter returns by DATE of return, regardless of when sale happened (Cash flow perspective)
    # Or by Sale Date? Usually "Top Products" for a period implies Net Movement in that period.
    returns_query = db.query(
        models.Product.id,
        func.sum(models.ReturnDetail.quantity).label('ret_qty'),
        func.sum(models.ReturnDetail.quantity * models.ReturnDetail.unit_price).label('ret_rev') # Approximate revenue reversal
    ).join(models.ReturnDetail).join(models.Return).filter(
        models.Return.date >= start_dt,
        models.Return.date <= end_dt
    ).group_by(models.Product.id).all()
    
    returns_map = {r[0]: {"qty": float(r[1]), "rev": float(r[2] or 0)} for r in returns_query}
    
    # 3. Calculate Net
    final_list = []
    for pid, data in sales_map.items():
        ret_data = returns_map.get(pid, {"qty": 0, "rev": 0})
        net_qty = data["gross_qty"] - ret_data["qty"]
        net_rev = data["gross_rev"] - ret_data["rev"]
        
        if net_qty > 0: # Only show positive net sales
            final_list.append({
                "product_id": pid,
                "product_name": data["name"],
                "quantity_sold": net_qty,
                "revenue": net_rev,
                "gross_quantity": data["gross_qty"],
                "returned_quantity": ret_data["qty"]
            })
            
    # 4. Sort and Limit
    if by == "quantity":
        final_list.sort(key=lambda x: x["quantity_sold"], reverse=True)
    else:
        final_list.sort(key=lambda x: x["revenue"], reverse=True)
        
    return final_list[:limit]

@router.get("/customer-debts")
def get_customer_debt_report(db: Session = Depends(get_db)):
    """All customers with outstanding debt"""
    customers = db.query(models.Customer).all()
    
    report = []
    for customer in customers:
        # Sum unpaid credit sales
        unpaid_sales = db.query(func.sum(models.Sale.total_amount)).filter(
            models.Sale.customer_id == customer.id,
            models.Sale.is_credit == True,
            models.Sale.paid == False
        ).scalar() or 0
        
        # Sum payments
        payments = db.query(func.sum(models.Payment.amount)).filter(
            models.Payment.customer_id == customer.id
        ).scalar() or 0
        
        debt = unpaid_sales - payments
        
        if debt > 0:
            report.append({
                "customer_id": customer.id,
                "customer_name": customer.name,
                "phone": customer.phone,
                "debt": debt
            })
    
    # Sort by debt descending
    report.sort(key=lambda x: x["debt"], reverse=True)
    return report

@router.get("/low-stock")
def get_low_stock_products(threshold: int = 5, db: Session = Depends(get_db)):
    """Products with stock <= threshold"""
    products = db.query(models.Product).filter(models.Product.stock <= threshold).all()
    return products

@router.get("/inventory-valuation")
def get_inventory_valuation(exchange_rate: float = 1.0, db: Session = Depends(get_db)):
    """
    Inventory Financials:
    - Total Cost: Sum(Stock * Cost Price)
    - Potential Revenue: Sum(Stock * Sale Price)
    - Potential Profit: Revenue - Cost
    """
    products = db.query(models.Product).filter(models.Product.is_active == True).all()
    
    total_cost_usd = Decimal("0.00")
    total_revenue_usd = Decimal("0.00")
    total_stock_units = Decimal("0.00")
    
    for p in products:
        # Safety check for None and force Decimal conversion
        stock = Decimal(str(p.stock)) if p.stock is not None else Decimal("0")
        cost = Decimal(str(p.cost_price)) if p.cost_price is not None else Decimal("0")
        price = Decimal(str(p.price)) if p.price is not None else Decimal("0")
        
        # Only count positive stock
        if stock > 0:
            total_stock_units += stock
            total_cost_usd += (stock * cost)
            total_revenue_usd += (stock * price)
            
    potential_profit = total_revenue_usd - total_cost_usd
    margin_percent = 0
    if total_revenue_usd > 0:
        margin_percent = (potential_profit / total_revenue_usd) * 100

    return {
        "total_products": len(products),
        "total_stock_units": float(total_stock_units),
        "total_cost_usd": float(total_cost_usd),
        "total_revenue_usd": float(total_revenue_usd),
        "potential_profit_usd": float(potential_profit),
        "margin_percent": float(round(margin_percent, 2)),
        # Bs values for display if needed
        "total_cost_bs": float(total_cost_usd) * exchange_rate,
        "total_revenue_bs": float(total_revenue_usd) * exchange_rate
    }

# ===== PROFIT ANALYSIS ENDPOINTS =====

@router.get("/profit/product/{product_id}")
def get_product_profitability(product_id: int, db: Session = Depends(get_db)):
    """Get profitability stats for a specific product"""
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Get total sold
    total_sold = db.query(func.sum(models.SaleDetail.quantity)).filter(
        models.SaleDetail.product_id == product_id
    ).scalar() or 0
    
    # Calculate total profit
    sales = db.query(models.SaleDetail).filter(models.SaleDetail.product_id == product_id).all()
    
    total_profit = 0
    # Update logic to use historical cost
    for detail in sales:
        cost_price = detail.cost_at_sale if (detail.cost_at_sale is not None and detail.cost_at_sale > 0) else product.cost_price
        total_profit += (detail.unit_price - cost_price) * detail.quantity
    
    margin = 0
    if product.price > 0:
        margin = ((product.price - product.cost_price) / product.price) * 100
    
    return {
        'product_id': product.id,
        'product_name': product.name,
        'cost': product.cost_price,
        'price': product.price,
        'margin_percent': margin,
        'total_sold': total_sold,
        'total_profit': total_profit
    }

@router.get("/profit/sales")
def get_sales_profitability(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """Get total profitability for a date range"""
    query = db.query(models.SaleDetail).join(models.Sale)
    
    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time())
        query = query.filter(models.Sale.date >= start_dt)
    if end_date:
        end_dt = datetime.combine(end_date, datetime.max.time())
        query = query.filter(models.Sale.date <= end_dt)
    
    details = query.all()
    
    total_revenue = 0
    total_cost = 0
    
    for detail in details:
        total_revenue += float(detail.subtotal)
        
        # HISTORICAL COST LOGIC
        cost_price = float(detail.cost_at_sale if (detail.cost_at_sale is not None and detail.cost_at_sale > 0) else detail.product.cost_price)
        total_cost += cost_price * float(detail.quantity)
        
    # --- SUBTRACT RETURNS (Net Financials) ---
    returns_query = db.query(models.ReturnDetail).join(models.Return).filter(
        models.Return.date >= (start_dt if start_date else datetime.min),
        models.Return.date <= (end_dt if end_date else datetime.max)
    ).options(joinedload(models.ReturnDetail.product)).all()
    
    total_refunds = 0
    total_refund_cost = 0
    
    for rd in returns_query:
        qty = float(rd.quantity)
        refund_amt = float(rd.unit_price) * qty
        total_refunds += refund_amt
        
        # Restock Cost Reversal (Using historical unit_cost from return detail)
        r_cost = float(rd.unit_cost if (rd.unit_cost is not None and rd.unit_cost > 0) else rd.product.cost_price)
        total_refund_cost += r_cost * qty

    # Calculate Nets
    net_revenue = total_revenue - total_refunds
    net_cost = total_cost - total_refund_cost
    net_profit = net_revenue - net_cost
    
    avg_margin = 0
    if net_revenue > 0:
        avg_margin = (net_profit / net_revenue) * 100
        
    # --- REALIZED PROFIT (Cash Basis Approximation) ---
    # Formula: Realized Profit = Total Profit * (Collection Ratio)
    # Collection Ratio = (Total Revenue - Total Outstanding) / Total Revenue
    
    # Calculate Total Outstanding Debt for these specific sales
    total_outstanding = 0
    sale_ids = set(d.sale_id for d in details)
    
    if sale_ids:
        # Fetch scales to get balance_pending
        sales_in_period = db.query(models.Sale).filter(models.Sale.id.in_(sale_ids)).all()
        for s in sales_in_period:
            if s.is_credit and not s.paid:
                 total_outstanding += float(s.balance_pending if s.balance_pending is not None else s.total_amount)
    
    # Realized Revenue (Cash In)
    realized_revenue = max(0, net_revenue - total_outstanding)
    
    # Collection Ratio
    collection_ratio = 1.0
    if net_revenue > 0:
        collection_ratio = realized_revenue / net_revenue
    
    realized_profit = net_profit * collection_ratio
    
    return {
        'total_revenue': net_revenue,
        'total_cost': net_cost,
        'total_profit': net_profit,
        'avg_margin': avg_margin,
        'gross_revenue': total_revenue,
        'returns_amount': total_refunds,
        'num_sales': len(sale_ids),
        'realized_profit': realized_profit,
        'outstanding_debt_in_period': total_outstanding,
        'collection_ratio': collection_ratio
    }

@router.get("/profit/month")
def get_month_profitability(db: Session = Depends(get_db)):
    """Get profitability for current month"""
    now = datetime.now()
    start_of_month = datetime(now.year, now.month, 1)
    
    query = db.query(models.SaleDetail).join(models.Sale).filter(
        models.Sale.date >= start_of_month,
        models.Sale.date <= now
    )
    
    details = query.all()
    
    total_revenue = 0
    total_cost = 0
    
    for detail in details:
        total_revenue += detail.subtotal
        
        # HISTORICAL COST LOGIC
        cost_price = detail.cost_at_sale if (detail.cost_at_sale is not None and detail.cost_at_sale > 0) else detail.product.cost_price
        total_cost += cost_price * detail.quantity
    
    total_profit = total_revenue - total_cost
    avg_margin = 0
    if total_revenue > 0:
        avg_margin = (total_profit / total_revenue) * 100
    
    return {
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'avg_margin': avg_margin,
        'num_sales': len(set(d.sale_id for d in details))
    }


@router.get("/daily-close")
def get_daily_close(
    date: date,
    db: Session = Depends(get_db)
):
    """
    Daily Closing Report:
    - Sales by Payment Method
    - Sales by Cashier (User)
    - Cash Flow Summary
    """
    start_dt = datetime.combine(date, datetime.min.time())
    end_dt = datetime.combine(date, datetime.max.time())
    
    # 1. Sales by Payment Method (DETAILED Breakdown from Payments Table)
    # This separates Cash USD from Cash Bs, and handles split payments correctly.
    sales_by_method = db.query(
        models.SalePayment.payment_method,
        models.SalePayment.currency,
        func.sum(models.SalePayment.amount).label('total'),
        func.count(models.SalePayment.id).label('count')
    ).join(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).group_by(
        models.SalePayment.payment_method,
        models.SalePayment.currency
    ).all()
    
    # Format for Frontend (e.g. "Efectivo (USD)", "Efectivo (Bs)")
    formatted_breakdown = []
    for r in sales_by_method:
        method = r[0] or "N/A"
        currency = r[1] or "USD"
        label = f"{method} ({currency})"
        
        formatted_breakdown.append({
            "method": label,
            "total": float(r[2]),
            "count": r[3],
            "raw_currency": currency # Helpful for frontend if needed
        })
    
    # 2. Total Change Given (Vueltos)
    total_change_query = db.query(
        func.sum(models.Sale.change_amount)
    ).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).scalar() or 0.00
    
    return {
        "date": date.isoformat(),
        "sales_by_method": formatted_breakdown,
        "total_change_given": float(total_change_query),
        "system_status": "OK"
    }


# ===== EXCEL EXPORT ENDPOINT =====
from ..services import sales_export_service
from fastapi.responses import StreamingResponse
import io

@router.get("/export/sales", summary="Exportar Ventas a Excel Detallado")
def export_sales_excel(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    # current_user: models.User = Depends(get_current_active_user) # Uncomment when auth is ready or if generic
):
    """
    Genera un archivo Excel (.xlsx) con el desglose financiero exacto:
    - Costos (Históricos)
    - Ganancia Real
    - Totales en USD y VES
    - Tasa Implícita
    """
    
    excel_file = sales_export_service.generate_sales_excel(db, start_date, end_date)
    
    # Filename with dates
    filename = f"Ventas_Detalladas_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




@router.get("/export/excel")
async def export_excel_report(
    start_date: Optional[date] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """
    Generate a comprehensive management report in Excel format.
    
    **Requires pandas and openpyxl:**
    ```bash
    pip install pandas openpyxl
    ```
    
    Returns a multi-sheet Excel file with:
    - **Dashboard**: Summary KPIs (Total Sales, Profit, Top 5 Products)
    - **Sales Detail**: All sales in the period
    - **Cash Audit**: Cash sessions with discrepancies highlighted
    - **Inventory**: Current inventory valuation
    """
    try:
        import pandas as pd
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="pandas and openpyxl are required. Install with: pip install pandas openpyxl"
        )
    
    # Set default dates if not provided (current month)
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = date(end_date.year, end_date.month, 1)
    
    try:
        # ============================================
        # 1. QUERY DATA FROM DATABASE
        # ============================================
        
        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time())
        
        # Query Sales
        sales_query = db.query(
            models.Sale.id,
            models.Sale.date,
            models.Sale.total_amount,
            models.Sale.payment_method,
            models.Sale.paid,
            models.Customer.name.label('customer_name')
        ).outerjoin(
            models.Customer, models.Sale.customer_id == models.Customer.id
        ).filter(
            models.Sale.date >= start_dt,
            models.Sale.date <= end_dt
        ).all()
        
        # Query Sale Details for product analysis
        sale_details_query = db.query(
            models.SaleDetail.product_id,
            models.Product.name.label('product_name'),
            models.Product.cost_price,
            func.sum(models.SaleDetail.quantity).label('total_quantity'),
            func.sum(models.SaleDetail.subtotal).label('total_revenue')
        ).join(
            models.Product, models.SaleDetail.product_id == models.Product.id
        ).join(
            models.Sale, models.SaleDetail.sale_id == models.Sale.id
        ).filter(
            models.Sale.date >= start_dt,
            models.Sale.date <= end_dt
        ).group_by(
            models.SaleDetail.product_id,
            models.Product.name,
            models.Product.cost_price
        ).all()
        
        # Query Cash Sessions
        cash_sessions_query = db.query(
            models.CashSession.id,
            models.CashSession.start_time,
            models.CashSession.end_time,
            models.CashSession.initial_cash,
            models.CashSession.final_cash_expected,
            models.CashSession.final_cash_reported,
            models.CashSession.status,
            models.User.full_name.label('cashier_name')
        ).outerjoin(
            models.User, models.CashSession.user_id == models.User.id
        ).filter(
            models.CashSession.start_time >= start_dt,
            models.CashSession.start_time <= end_dt
        ).all()
        
        # Query Current Inventory
        inventory_query = db.query(
            models.Product.id,
            models.Product.name,
            models.Product.sku,
            models.Product.stock,
            models.Product.cost_price,
            models.Product.price,
            models.Category.name.label('category_name')
        ).outerjoin(
            models.Category, models.Product.category_id == models.Category.id
        ).filter(
            models.Product.is_active == True
        ).all()
        
        # ============================================
        # 2. CONVERT TO PANDAS DATAFRAMES
        # ============================================
        
        # Sales DataFrame
        sales_data = [{
            'ID Venta': s.id,
            'Fecha': s.date.strftime('%Y-%m-%d %H:%M') if s.date else '',
            'Cliente': s.customer_name or 'Público General',
            'Total': float(s.total_amount or 0),
            'Método Pago': s.payment_method or 'N/A',
            'Pagado': 'Sí' if s.paid else 'No'
        } for s in sales_query]
        df_sales = pd.DataFrame(sales_data)
        
        # Product Sales DataFrame
        product_sales_data = [{
            'Producto': p.product_name,
            'Cantidad Vendida': float(p.total_quantity or 0),
            'Ingresos Totales': float(p.total_revenue or 0),
            'Costo Unitario': float(p.cost_price or 0),
            'Ganancia Estimada': float(p.total_revenue or 0) - (float(p.cost_price or 0) * float(p.total_quantity or 0))
        } for p in sale_details_query]
        df_products = pd.DataFrame(product_sales_data)
        
        # Cash Sessions DataFrame
        cash_data = [{
            'ID Sesión': c.id,
            'Cajero': c.cashier_name or f'Usuario #{c.id}',
            'Apertura': c.start_time.strftime('%Y-%m-%d %H:%M') if c.start_time else '',
            'Cierre': c.end_time.strftime('%Y-%m-%d %H:%M') if c.end_time else 'Abierta',
            'Inicial': float(c.initial_cash or 0),
            'Esperado': float(c.final_cash_expected or 0),
            'Reportado': float(c.final_cash_reported or 0),
            'Diferencia': float(c.final_cash_reported or 0) - float(c.final_cash_expected or 0),
            'Estado': c.status
        } for c in cash_sessions_query]
        df_cash = pd.DataFrame(cash_data)
        
        # Inventory DataFrame
        inventory_data = [{
            'SKU': i.sku or '',
            'Producto': i.name,
            'Categoría': i.category_name or 'Sin categoría',
            'Stock': float(i.stock or 0),
            'Costo': float(i.cost_price or 0),
            'Precio': float(i.price or 0),
            'Valor Inventario': float(i.stock or 0) * float(i.cost_price or 0)
        } for i in inventory_query]
        df_inventory = pd.DataFrame(inventory_data)
        
        # ============================================
        # 3. CREATE EXCEL FILE WITH MULTIPLE SHEETS
        # ============================================
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # ========== SHEET 1: DASHBOARD ==========
            dashboard_data = {
                'Métrica': [
                    'Total Ventas',
                    'Ganancia Estimada',
                    'Número de Ventas',
                    'Ticket Promedio',
                    'Total Faltantes Caja',
                    'Total Sobrantes Caja',
                    'Valor Total Inventario'
                ],
                'Valor': [
                    f"${df_sales['Total'].sum():,.2f}" if not df_sales.empty else '$0.00',
                    f"${df_products['Ganancia Estimada'].sum():,.2f}" if not df_products.empty else '$0.00',
                    len(df_sales),
                    f"${df_sales['Total'].mean():,.2f}" if not df_sales.empty else '$0.00',
                    f"${df_cash[df_cash['Diferencia'] < 0]['Diferencia'].sum():,.2f}" if not df_cash.empty else '$0.00',
                    f"${df_cash[df_cash['Diferencia'] > 0]['Diferencia'].sum():,.2f}" if not df_cash.empty else '$0.00',
                    f"${df_inventory['Valor Inventario'].sum():,.2f}" if not df_inventory.empty else '$0.00'
                ]
            }
            df_dashboard = pd.DataFrame(dashboard_data)
            df_dashboard.to_excel(writer, sheet_name='Dashboard', index=False)
            
            # Add Top 5 Products
            if not df_products.empty:
                top_products = df_products.nlargest(5, 'Ingresos Totales')[['Producto', 'Cantidad Vendida', 'Ingresos Totales', 'Ganancia Estimada']]
                top_products.to_excel(writer, sheet_name='Dashboard', startrow=len(df_dashboard) + 3, index=False)
                
                # Add header for top products
                worksheet = writer.sheets['Dashboard']
                worksheet.cell(row=len(df_dashboard) + 3, column=1, value='TOP 5 PRODUCTOS MÁS VENDIDOS')
                worksheet.cell(row=len(df_dashboard) + 3, column=1).font = Font(bold=True, size=12)
            
            # ========== SHEET 2: SALES DETAIL ==========
            if not df_sales.empty:
                df_sales.to_excel(writer, sheet_name='Ventas Detalle', index=False)
            else:
                pd.DataFrame({'Mensaje': ['No hay ventas en este período']}).to_excel(writer, sheet_name='Ventas Detalle', index=False)
            
            # ========== SHEET 3: CASH AUDIT ==========
            if not df_cash.empty:
                df_cash.to_excel(writer, sheet_name='Auditoría Cajas', index=False)
            else:
                pd.DataFrame({'Mensaje': ['No hay sesiones de caja en este período']}).to_excel(writer, sheet_name='Auditoría Cajas', index=False)
            
            # ========== SHEET 4: INVENTORY ==========
            if not df_inventory.empty:
                df_inventory.to_excel(writer, sheet_name='Inventario', index=False)
            else:
                pd.DataFrame({'Mensaje': ['No hay productos en inventario']}).to_excel(writer, sheet_name='Inventario', index=False)
        
        # ============================================
        # 4. APPLY STYLING TO EXCEL
        # ============================================
        
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        
        # Style Dashboard
        if 'Dashboard' in workbook.sheetnames:
            ws = workbook['Dashboard']
            
            # Header styling
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Style Cash Audit - Highlight differences
        if 'Auditoría Cajas' in workbook.sheetnames:
            ws = workbook['Auditoría Cajas']
            
            # Header styling
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Highlight rows with differences
            diff_col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Diferencia':
                    diff_col_idx = idx
                    break
            
            if diff_col_idx:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    diff_value = row[diff_col_idx - 1].value
                    if diff_value and isinstance(diff_value, (int, float)):
                        if diff_value < -0.01:  # Shortage
                            for cell in row:
                                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        elif diff_value > 0.01:  # Overage
                            for cell in row:
                                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        
        # Save styled workbook
        final_output = BytesIO()
        workbook.save(final_output)
        final_output.seek(0)
        
        # ============================================
        # 5. RETURN FILE AS DOWNLOAD
        # ============================================
        
        filename = f"Reporte_Gerencial_{start_date}_{end_date}.xlsx"
        
        return StreamingResponse(
            final_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")


@router.get("/export/general")
async def export_general_report(
    start_date: Optional[date] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """
    Generate a comprehensive 360° audit report in Excel format with flattened multi-currency columns.
    
    Returns a multi-sheet Excel file with:
    - **Dashboard**: Summary KPIs
    - **Auditoría de Cajas**: FLATTENED multi-currency columns (USD Reportado, Dif USD, BS Reportado, Dif BS, etc.)
    - **Ventas Detalladas**: All sales in the period
    """
    try:
        import pandas as pd
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="pandas and openpyxl are required. Install with: pip install pandas openpyxl"
        )
    
    # Set default dates if not provided (current month)
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = date(end_date.year, end_date.month, 1)
    
    try:
        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time())
        
        # ============================================
        # 1. QUERY CASH SESSIONS WITH CURRENCIES
        # ============================================
        
        cash_sessions = db.query(models.CashSession).options(
            joinedload(models.CashSession.currencies),
            joinedload(models.CashSession.user)
        ).filter(
            models.CashSession.start_time >= start_dt,
            models.CashSession.start_time <= end_dt
        ).all()
        
        # ============================================
        # 2. FLATTEN MULTI-CURRENCY DATA
        # ============================================
        
        audit_data = []
        
        for session in cash_sessions:
            row = {
                'Fecha Apertura': session.start_time.strftime('%Y-%m-%d %H:%M') if session.start_time else '',
                'Fecha Cierre': session.end_time.strftime('%Y-%m-%d %H:%M') if session.end_time else 'Abierta',
                'Cajero': session.user.full_name if session.user and session.user.full_name else (session.user.username if session.user else f'Usuario #{session.user_id}'),
                'Estado': session.status,
            }
            
            # Process currencies - flatten into separate columns
            if session.currencies and len(session.currencies) > 0:
                for curr in session.currencies:
                    symbol = curr.currency_symbol
                    row[f'{symbol} Inicial'] = float(curr.initial_amount or 0)
                    row[f'{symbol} Esperado'] = float(curr.final_expected or 0)
                    row[f'{symbol} Reportado'] = float(curr.final_reported or 0)
                    row[f'Dif {symbol}'] = float(curr.difference or 0)
            else:
                # Legacy session - only USD
                row['USD Inicial'] = float(session.initial_cash or 0)
                row['USD Esperado'] = float(session.final_cash_expected or 0)
                row['USD Reportado'] = float(session.final_cash_reported or 0)
                diff = float(session.final_cash_reported or 0) - float(session.final_cash_expected or 0)
                row['Dif USD'] = diff
            
            audit_data.append(row)
        
        df_audit = pd.DataFrame(audit_data) if audit_data else pd.DataFrame()
        
        # ============================================
        # 3. QUERY SALES DATA (SIMPLIFIED)
        # ============================================
        
        sales_query = db.query(
            models.Sale.id,
            models.Sale.date,
            models.Customer.name.label('customer_name'),
            models.Sale.total_amount,
            models.Sale.payment_method
        ).outerjoin(
            models.Customer, models.Sale.customer_id == models.Customer.id
        ).filter(
            models.Sale.date >= start_dt,
            models.Sale.date <= end_dt
        ).all()
        
        sales_data = [{
            'ID': s.id,
            'Fecha': s.date.strftime('%Y-%m-%d %H:%M') if s.date else '',
            'Cliente': s.customer_name or 'Público General',
            'Total': float(s.total_amount or 0),
            'Método Pago': s.payment_method or 'N/A'
        } for s in sales_query]
        df_sales = pd.DataFrame(sales_data) if sales_data else pd.DataFrame()
        
        # ============================================
        # 4. CALCULATE DASHBOARD KPIS
        # ============================================
        
        total_sales = df_sales['Total'].sum() if not df_sales.empty else 0
        num_sales = len(df_sales)
        avg_ticket = total_sales / num_sales if num_sales > 0 else 0
        
        # Calculate shortages and overages from audit data
        total_shortages = 0
        total_overages = 0
        
        if not df_audit.empty:
            for col in df_audit.columns:
                if col.startswith('Dif '):
                    for val in df_audit[col]:
                        if pd.notna(val):
                            if val < -0.01:
                                total_shortages += abs(val)
                            elif val > 0.01:
                                total_overages += val
        
        dashboard_data = {
            'Métrica': [
                'Total Ventas USD',
                'Número de Ventas',
                'Ticket Promedio',
                'Total Faltantes',
                'Total Sobrantes',
                'Sesiones Auditadas'
            ],
            'Valor': [
                f"${total_sales:,.2f}",
                num_sales,
                f"${avg_ticket:,.2f}",
                f"${total_shortages:,.2f}",
                f"${total_overages:,.2f}",
                len(cash_sessions)
            ]
        }
        df_dashboard = pd.DataFrame(dashboard_data)
        
        # ============================================
        # 5. CREATE EXCEL FILE
        # ============================================
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Dashboard
            df_dashboard.to_excel(writer, sheet_name='Dashboard', index=False)
            
            # Sheet 2: Auditoría de Cajas (FLATTENED)
            if not df_audit.empty:
                df_audit.to_excel(writer, sheet_name='Auditoría de Cajas', index=False)
            else:
                pd.DataFrame({'Mensaje': ['No hay sesiones de caja en este período']}).to_excel(
                    writer, sheet_name='Auditoría de Cajas', index=False
                )
            
            # Sheet 3: Ventas Detalladas
            if not df_sales.empty:
                df_sales.to_excel(writer, sheet_name='Ventas Detalladas', index=False)
            else:
                pd.DataFrame({'Mensaje': ['No hay ventas en este período']}).to_excel(
                    writer, sheet_name='Ventas Detalladas', index=False
                )
        
        # ============================================
        # 6. APPLY STYLING
        # ============================================
        
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        
        # Style Dashboard
        if 'Dashboard' in workbook.sheetnames:
            ws = workbook['Dashboard']
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style Auditoría de Cajas - Highlight differences
        if 'Auditoría de Cajas' in workbook.sheetnames:
            ws = workbook['Auditoría de Cajas']
            
            # Header styling
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Find all "Dif" columns and highlight
            diff_cols = []
            for idx, cell in enumerate(ws[1], 1):
                if cell.value and str(cell.value).startswith('Dif '):
                    diff_cols.append(idx)
            
            # Apply conditional formatting to difference columns
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in diff_cols:
                    cell = row[col_idx - 1]
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value < -0.01:  # Shortage
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                            cell.font = Font(bold=True, color="CC0000")
                        elif cell.value > 0.01:  # Overage
                            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                            cell.font = Font(bold=True, color="00CC00")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save styled workbook
        final_output = BytesIO()
        workbook.save(final_output)
        final_output.seek(0)
        
        # ============================================
        # 7. RETURN FILE
        # ============================================
        
        filename = f"Auditoria_360_General_{start_date}_{end_date}.xlsx"
        
        return StreamingResponse(
            final_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating general report: {str(e)}")



# ===== DETAILED REPORTS =====

@router.get("/sales/by-payment-method")
def get_sales_by_payment_method(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db)
):
    """
    Sales breakdown by payment method.
    Groups by 'payment_method' column in Sale.
    """
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # Query: Group by payment_method, sum total_amount
    results = db.query(
        models.Sale.payment_method,
        func.sum(models.Sale.total_amount).label('total_amount'),
        func.count(models.Sale.id).label('count')
    ).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
        # Note: We include all statuses or filter by active? 
        # Ideally only COMPLETED for reports
    ).group_by(models.Sale.payment_method).all()
    
    # Filter COMPLETED only roughly (re-using logic from sales search)
    # Better: join returns to exclude voided, or trust the data if we have status column
    # Since we don't have a physical status column in DB yet (it's property), 
    # and doing a join here is expensive for aggregation without proper indexing,
    # let's proceed with simple aggregation but filter out known "VOIDED" if we can via existing logic.
    # Actually, let's keep it simple: Raw sales data.
    
    return [
        {
            "method": r.payment_method or "Desconocido",
            "total_amount": float(r.total_amount or 0),
            "count": r.count
        }
        for r in results
    ]

@router.get("/sales/by-customer")
def get_sales_by_customer(
    start_date: date,
    end_date: date,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """
    Top customers by sales volume.
    """
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    results = db.query(
        models.Customer.name,
        func.sum(models.Sale.total_amount).label('total_purchased'),
        func.count(models.Sale.id).label('transaction_count')
    ).join(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).group_by(models.Customer.id, models.Customer.name)\
    .order_by(desc('total_purchased'))\
    .limit(limit).all()
    
    return [
        {
            "customer_name": r.name,
            "total_purchased": float(r.total_purchased or 0),
            "transaction_count": r.transaction_count
        }
        for r in results
    ]
@router.get("/export/detailed")
def export_detailed_report(
    start_date: date,
    end_date: date,
    format: str = "xlsx",
    db: Session = Depends(get_db)
):
    """
    Export detailed combined report to Excel (Multi-sheet).
    Sheet 1: Payment Methods
    Sheet 2: Top Customers
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        workbook = Workbook()
        # Remove default sheet
        default_ws = workbook.active
        workbook.remove(default_ws)

        def create_sheet(wb, sheet_title, report_title, headers, data):
            ws = wb.create_sheet(title=sheet_title)
            
            # Title Row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.cell(row=1, column=1, value=f"{report_title} ({start_date} - {end_date})").font = Font(size=14, bold=True)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            # Header Row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
            # Data Rows
            row_idx = 4
            for row_data in data:
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    # Format currency columns (usually last column)
                    if col_idx == 3: 
                         cell.number_format = '$#,##0.00'
                row_idx += 1
                
            # Total Row
            if data:
                ws.cell(row=row_idx, column=1, value="TOTAL GENERADO").font = Font(bold=True)
                total_sales = sum(row[2] for row in data)
                total_cell = ws.cell(row=row_idx, column=3, value=total_sales)
                total_cell.font = Font(bold=True)
                total_cell.number_format = '$#,##0.00'
                
            # Auto-adjust column width
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

        # --- SHEET 1: Payment Methods ---
        raw_payments = get_sales_by_payment_method(start_date, end_date, db)
        data_payments = [[r['method'], r['count'], r['total_amount']] for r in raw_payments]
        create_sheet(workbook, "Métodos de Pago", "Ventas por Método de Pago", ["Método", "Transacciones", "Total (USD)"], data_payments)

        # --- SHEET 2: Customers ---
        raw_customers = get_sales_by_customer(start_date, end_date, limit=100, db=db)
        data_customers = [[r['customer_name'], r['transaction_count'], r['total_purchased']] for r in raw_customers]
        create_sheet(workbook, "Clientes Top", "Ventas por Cliente", ["Cliente", "Compras", "Total (USD)"], data_customers)

        # Save & Return
        final_output = BytesIO()
        workbook.save(final_output)
        final_output.seek(0)
        
        filename = f"Reporte_Completo_{start_date}_{end_date}.xlsx"
        
        return StreamingResponse(
            final_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting report: {str(e)}")
