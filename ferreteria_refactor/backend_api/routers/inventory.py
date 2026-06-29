from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, or_
from typing import List, Optional, Dict, Any
from pydantic import BaseModel
from ..database.db import get_db
from ..models import models
from .. import schemas
from datetime import datetime, timedelta
from ..dependencies import warehouse_or_admin, require_permission, require_any_permission, get_current_active_user
from ..websocket.manager import manager
from ..websocket.events import WebSocketEvents
from ..tenant_context import get_tenant_schema

router = APIRouter(
    prefix="/inventory",
    tags=["inventory"],
    dependencies=[]  # Dependencies moved to individual endpoints
)

@router.post("/add")
async def add_stock(adjustment: schemas.StockAdjustmentCreate, db: Session = Depends(get_db), current_user: models.User = Depends(warehouse_or_admin)):
    """Add stock (Purchase/Entry) - Multi-Warehouse Support"""
    product = db.query(models.Product).filter(models.Product.id == adjustment.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # 1. Update/Create Specific Warehouse Stock
    product_stock = db.query(models.ProductStock).filter(
        models.ProductStock.product_id == product.id,
        models.ProductStock.warehouse_id == adjustment.warehouse_id
    ).first()

    if not product_stock:
        # Check if warehouse exists first
        warehouse = db.query(models.Warehouse).filter(models.Warehouse.id == adjustment.warehouse_id).first()
        if not warehouse:
             raise HTTPException(status_code=404, detail="Warehouse not found")

        product_stock = models.ProductStock(
            product_id=product.id,
            warehouse_id=adjustment.warehouse_id,
            quantity=0
        )
        db.add(product_stock)
    
    product_stock.quantity += adjustment.quantity

    # 2. Update Global Stock (Cache)
    product.stock += adjustment.quantity
    
    # Create Kardex
    kardex_entry = models.Kardex(
        product_id=product.id,
        warehouse_id=adjustment.warehouse_id, # NEW
        movement_type=adjustment.type,
        quantity=adjustment.quantity,
        balance_after=product.stock, # Keeping global balance for now
        description=adjustment.reason,
        date=datetime.now()
    )
    
    db.add(kardex_entry)

    # Capture state BEFORE commit to avoid ObjectDeletedError/SessionExpire
    product_id = product.id
    product_name = product.name
    product_price = float(product.price)
    product_stock = float(product.stock)
    product_er_id = product.exchange_rate_id

    db.commit()
    # db.refresh(product) <- REMOVED
    
    # AUDIT LOG
    from ..audit_utils import log_action
    log_action(db, user_id=current_user.id, action="UPDATE", table_name="products", record_id=product_id, changes=f"Stock Adjustment (IN) [Wh:{adjustment.warehouse_id}]: +{adjustment.quantity}. Reason: {adjustment.reason}")

    tenant_schema = get_tenant_schema()
    await manager.broadcast(WebSocketEvents.PRODUCT_UPDATED, {
        "id": product_id,
        "name": product_name,
        "price": product_price,
        "stock": product_stock,
        "exchange_rate_id": product_er_id
    }, tenant_id=tenant_schema)
    
    await manager.broadcast(WebSocketEvents.PRODUCT_STOCK_UPDATED, {
        "id": product_id,
        "stock": product_stock
    }, tenant_id=tenant_schema)
    
    return {"status": "success", "new_stock": product_stock, "product_id": product_id}

@router.post("/remove")
async def remove_stock(adjustment: schemas.StockAdjustmentCreate, db: Session = Depends(get_db), current_user: models.User = Depends(warehouse_or_admin)):
    """Remove stock (Adjustment/Loss) - Multi-Warehouse Support"""
    product = db.query(models.Product).filter(models.Product.id == adjustment.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # 1. Validate Specific Warehouse Stock
    product_stock = db.query(models.ProductStock).filter(
        models.ProductStock.product_id == product.id,
        models.ProductStock.warehouse_id == adjustment.warehouse_id
    ).first()

    if not product_stock:
         raise HTTPException(status_code=400, detail="Producto no tiene stock en esta bodega")
    
    if product_stock.quantity < adjustment.quantity:
        raise HTTPException(status_code=400, detail=f"Stock insuficiente en bodega (Disponible: {product_stock.quantity})")

    # 2. Update Specific Stock
    product_stock.quantity -= adjustment.quantity

    # 3. Update Global Stock (Cache)
    product.stock -= adjustment.quantity
    
    # Create Kardex
    
    # Map Frontend Types to DB Enum
    db_movement_type = adjustment.type
    final_description = adjustment.reason
    
    if adjustment.type == "DAMAGED":
        db_movement_type = "ADJUSTMENT_OUT" # Or "ADJUSTMENT" based on models.py comment, but OUT is safer for logic
        final_description = f"[DAMAGED] {adjustment.reason}"
    elif adjustment.type == "INTERNAL_USE":
        db_movement_type = "ADJUSTMENT_OUT"
        final_description = f"[INTERNAL USE] {adjustment.reason}"
        
    kardex_entry = models.Kardex(
        product_id=product.id,
        warehouse_id=adjustment.warehouse_id, # NEW
        movement_type=db_movement_type,
        quantity=-adjustment.quantity,  # Negative for outgoing
        balance_after=product.stock,
        description=final_description,
        date=datetime.now()
    )
    
    db.add(kardex_entry)

    # Capture state BEFORE commit
    product_id = product.id
    product_name = product.name
    product_price = float(product.price)
    product_stock = float(product.stock)
    product_er_id = product.exchange_rate_id

    db.commit()
    # db.refresh(product) <- REMOVED
    
    # AUDIT LOG
    from ..audit_utils import log_action
    log_action(db, user_id=current_user.id, action="UPDATE", table_name="products", record_id=product_id, changes=f"Stock Adjustment (OUT) [Wh:{adjustment.warehouse_id}]: -{adjustment.quantity}. Reason: {adjustment.reason}")

    tenant_schema = get_tenant_schema()
    await manager.broadcast(WebSocketEvents.PRODUCT_UPDATED, {
        "id": product_id,
        "name": product_name,
        "price": product_price,
        "stock": product_stock,
        "exchange_rate_id": product_er_id
    }, tenant_id=tenant_schema)
    
    await manager.broadcast(WebSocketEvents.PRODUCT_STOCK_UPDATED, {
        "id": product_id,
        "stock": product_stock
    }, tenant_id=tenant_schema)
    
    return {"status": "success", "new_stock": product_stock, "product_id": product_id}


@router.get("/kardex", response_model=List[schemas.KardexRead], dependencies=[Depends(require_permission("inventory.kardex.view"))])
def get_kardex(
    product_id: Optional[int] = None, 
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 100, 
    db: Session = Depends(get_db)
):
    from sqlalchemy.orm import joinedload
    query = db.query(models.Kardex).options(joinedload(models.Kardex.product))
    
    if product_id:
        query = query.filter(models.Kardex.product_id == product_id)
        
    if start_date:
        query = query.filter(models.Kardex.date >= start_date)
        
    if end_date:
        # Include the whole end day by adding time or next day logic if needed. 
        # Assuming format YYYY-MM-DD, strict comparison might miss same-day events if not handled.
        # Simple string compare works if client sends 'YYYY-MM-DD' and DB has 'YYYY-MM-DD HH:MM:SS'
        # To be inclusive of the end date, we generally want <= end_date + " 23:59:59" or < next_day
        query = query.filter(models.Kardex.date <= f"{end_date} 23:59:59")
        
    return query.order_by(models.Kardex.date.desc()).limit(limit).all()


class InventoryExportRequest(BaseModel):
    export_type: str = "catalog_basic"
    columns: List[str] = []
    search: Optional[str] = None
    category_id: Optional[int] = None
    warehouse_id: Optional[int] = None
    stock_filter: Optional[str] = None
    movement_type: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    include_inactive: bool = False
    include_price_lists: bool = False
    format: str = "xlsx"
    limit: int = 5000


EXPORT_PERMISSION_BY_TYPE = {
    "catalog_basic": "inventory.products.view",
    "catalog_prices": "inventory.products.view",
    "stock": "inventory.products.view",
    "prices": "inventory.products.view",
    "kardex": "inventory.kardex.view",
    "serials": "inventory.serials.view",
}

PRODUCT_EXPORT_COLUMNS = {
    "name": "Producto",
    "sku": "SKU / Codigo",
    "category": "Categoria",
    "stock": "Stock total",
    "warehouse": "Almacen",
    "warehouse_stock": "Stock almacen",
    "min_stock": "Stock minimo",
    "status": "Estado stock",
    "cost_price": "Costo",
    "price": "Precio venta",
    "profit_margin": "Margen %",
    "tax_rate": "IVA %",
    "supplier": "Proveedor",
    "location": "Ubicacion",
    "type": "Tipo",
    "description": "Descripcion",
}

KARDEX_EXPORT_COLUMNS = {
    "date": "Fecha",
    "product": "Producto",
    "sku": "SKU / Codigo",
    "movement_type": "Movimiento",
    "quantity": "Cantidad",
    "balance_after": "Saldo despues",
    "warehouse": "Almacen",
    "description": "Descripcion",
}

SERIAL_EXPORT_COLUMNS = {
    "product": "Producto",
    "sku": "SKU / Codigo",
    "serial_number": "Serial / IMEI",
    "status": "Estado",
    "color_name": "Color",
    "color_hex": "Color HEX",
    "warehouse": "Almacen",
    "cost": "Costo",
    "created_at": "Fecha recepcion",
}

DEFAULT_EXPORT_COLUMNS = {
    "catalog_basic": ["name", "sku", "category", "stock", "status"],
    "catalog_prices": ["name", "sku", "category", "price", "cost_price", "stock"],
    "stock": ["name", "sku", "category", "warehouse", "warehouse_stock", "stock", "min_stock", "status"],
    "prices": ["name", "sku", "category", "price", "cost_price", "profit_margin", "tax_rate"],
    "kardex": ["date", "product", "sku", "movement_type", "quantity", "balance_after", "warehouse", "description"],
    "serials": ["product", "sku", "serial_number", "status", "color_name", "warehouse", "created_at"],
}

MOVEMENT_LABELS_EXPORT = {
    "SALE": "Venta",
    "SALE_MODIFIER": "Extra de venta",
    "SALE_REVERSED": "Venta anulada",
    "PURCHASE": "Compra",
    "ADJUSTMENT": "Ajuste",
    "ADJUSTMENT_IN": "Ajuste de entrada",
    "ADJUSTMENT_OUT": "Ajuste de salida",
    "DAMAGED": "Danado",
    "INTERNAL_USE": "Uso interno",
    "RETURN": "Devolucion",
    "OUT": "Salida",
    "TRANSFER_IN": "Traslado recibido",
    "TRANSFER_OUT": "Traslado enviado",
    "EXTERNAL_TRANSFER_IN": "Traslado externo recibido",
    "EXTERNAL_TRANSFER_OUT": "Traslado externo enviado",
}


def _export_enum(value):
    if value is None:
        return ""
    return getattr(value, "value", str(value))


def _export_number(value):
    if value is None:
        return 0
    try:
        return float(value)
    except Exception:
        return 0


def _export_date(value):
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(value)


def _parse_export_date(value, end=False):
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
        if end and len(str(value)) <= 10:
            parsed = parsed + timedelta(days=1)
        return parsed
    except Exception:
        return None


def _stock_status(stock, min_stock):
    stock_value = _export_number(stock)
    min_value = _export_number(min_stock) or 5
    if stock_value <= 0:
        return "Agotado"
    if stock_value <= min_value:
        return "Bajo"
    return "Disponible"


def _product_type_label(product):
    flags = []
    if getattr(product, "is_service", False):
        flags.append("Servicio")
    if getattr(product, "has_imei", False):
        flags.append("Serial/IMEI")
    if getattr(product, "is_combo", False):
        flags.append("Combo")
    return ", ".join(flags) or "Producto"


def _apply_product_export_filters(query, payload: InventoryExportRequest):
    if not payload.include_inactive:
        query = query.filter(models.Product.is_active == True)
    if payload.search:
        tokens = [token for token in payload.search.strip().split() if token]
        for token in tokens:
            like = f"%{token}%"
            query = query.filter(or_(models.Product.name.ilike(like), models.Product.sku.ilike(like)))
    if payload.category_id:
        query = query.filter(models.Product.category_id == payload.category_id)
    if payload.stock_filter == "out_of_stock":
        query = query.filter(models.Product.stock <= 0)
    elif payload.stock_filter == "low_stock":
        query = query.filter(models.Product.stock > 0, models.Product.stock < func.coalesce(models.Product.min_stock, 5))
    elif payload.stock_filter == "in_stock":
        query = query.filter(models.Product.stock > 0)
    return query


def _load_price_list_maps(db: Session, product_ids: List[int]):
    price_lists = db.query(models.PriceList).filter(models.PriceList.is_active == True).order_by(models.PriceList.name).all()
    prices_by_product = {}
    if product_ids and price_lists:
        rows = db.query(models.ProductPrice).filter(models.ProductPrice.product_id.in_(product_ids)).all()
        for row in rows:
            prices_by_product.setdefault(row.product_id, {})[row.price_list_id] = row.price
    return price_lists, prices_by_product


def _build_product_row(product, columns, price_lists=None, prices_by_product=None, warehouse=None, warehouse_stock=None):
    price_lists = price_lists or []
    prices_by_product = prices_by_product or {}
    values = {
        "name": product.name,
        "sku": product.sku or "",
        "category": product.category.name if product.category else "Sin categoria",
        "stock": _export_number(product.stock),
        "warehouse": warehouse.name if warehouse else "Consolidado",
        "warehouse_stock": _export_number(warehouse_stock if warehouse_stock is not None else product.stock),
        "min_stock": _export_number(product.min_stock),
        "status": _stock_status(product.stock, product.min_stock),
        "cost_price": _export_number(product.cost_price),
        "price": _export_number(product.price),
        "profit_margin": _export_number(product.profit_margin),
        "tax_rate": _export_number(product.tax_rate),
        "supplier": product.supplier.name if getattr(product, "supplier", None) else "",
        "location": product.location or "",
        "type": _product_type_label(product),
        "description": product.description or "",
    }
    row = {PRODUCT_EXPORT_COLUMNS[col]: values.get(col, "") for col in columns if col in PRODUCT_EXPORT_COLUMNS}
    product_prices = prices_by_product.get(product.id, {})
    for price_list in price_lists:
        row[f"Lista: {price_list.name}"] = _export_number(product_prices.get(price_list.id, 0))
    return row


def _export_catalog_rows(db: Session, payload: InventoryExportRequest, columns: List[str], include_price_lists=False):
    query = db.query(models.Product).options(
        joinedload(models.Product.category),
        joinedload(models.Product.supplier),
    )
    query = _apply_product_export_filters(query, payload)
    products = query.order_by(func.lower(models.Product.name)).limit(min(payload.limit or 5000, 20000)).all()
    price_lists, prices_by_product = _load_price_list_maps(db, [p.id for p in products]) if include_price_lists else ([], {})
    return [_build_product_row(p, columns, price_lists, prices_by_product) for p in products], len(products)


def _export_stock_rows(db: Session, payload: InventoryExportRequest, columns: List[str]):
    stock_query = db.query(models.ProductStock).join(models.Product).join(models.Warehouse).options(
        joinedload(models.ProductStock.product).joinedload(models.Product.category),
        joinedload(models.ProductStock.product).joinedload(models.Product.supplier),
        joinedload(models.ProductStock.warehouse),
    )
    if not payload.include_inactive:
        stock_query = stock_query.filter(models.Product.is_active == True)
    if payload.search:
        tokens = [token for token in payload.search.strip().split() if token]
        for token in tokens:
            like = f"%{token}%"
            stock_query = stock_query.filter(or_(models.Product.name.ilike(like), models.Product.sku.ilike(like)))
    if payload.category_id:
        stock_query = stock_query.filter(models.Product.category_id == payload.category_id)
    if payload.warehouse_id:
        stock_query = stock_query.filter(models.ProductStock.warehouse_id == payload.warehouse_id)
    if payload.stock_filter == "out_of_stock":
        stock_query = stock_query.filter(models.ProductStock.quantity <= 0)
    elif payload.stock_filter == "low_stock":
        stock_query = stock_query.filter(models.ProductStock.quantity > 0, models.ProductStock.quantity < func.coalesce(models.Product.min_stock, 5))
    elif payload.stock_filter == "in_stock":
        stock_query = stock_query.filter(models.ProductStock.quantity > 0)

    limit = min(payload.limit or 5000, 20000)
    stock_rows = stock_query.order_by(models.Warehouse.name, func.lower(models.Product.name)).limit(limit).all()
    rows = [
        _build_product_row(item.product, columns, warehouse=item.warehouse, warehouse_stock=item.quantity)
        for item in stock_rows
    ]

    if not payload.warehouse_id and len(rows) < limit:
        product_ids_with_stock = [item.product_id for item in stock_rows]
        product_query = db.query(models.Product).options(
            joinedload(models.Product.category),
            joinedload(models.Product.supplier),
        )
        product_query = _apply_product_export_filters(product_query, payload)
        if product_ids_with_stock:
            product_query = product_query.filter(~models.Product.id.in_(product_ids_with_stock))
        remaining = max(limit - len(rows), 0)
        products_without_rows = product_query.order_by(func.lower(models.Product.name)).limit(remaining).all()
        rows.extend(_build_product_row(product, columns) for product in products_without_rows)

    return rows, len(rows)


def _export_kardex_rows(db: Session, payload: InventoryExportRequest, columns: List[str]):
    query = db.query(models.Kardex, models.Product, models.Warehouse).join(
        models.Product, models.Kardex.product_id == models.Product.id
    ).outerjoin(models.Warehouse, models.Kardex.warehouse_id == models.Warehouse.id)
    start_dt = _parse_export_date(payload.start_date)
    end_dt = _parse_export_date(payload.end_date, end=True)
    if start_dt:
        query = query.filter(models.Kardex.date >= start_dt)
    if end_dt:
        query = query.filter(models.Kardex.date < end_dt)
    if payload.search:
        tokens = [token for token in payload.search.strip().split() if token]
        for token in tokens:
            like = f"%{token}%"
            query = query.filter(or_(models.Product.name.ilike(like), models.Product.sku.ilike(like), models.Kardex.description.ilike(like)))
    if payload.category_id:
        query = query.filter(models.Product.category_id == payload.category_id)
    if payload.warehouse_id:
        query = query.filter(models.Kardex.warehouse_id == payload.warehouse_id)
    if payload.movement_type and payload.movement_type != "ALL":
        query = query.filter(models.Kardex.movement_type == payload.movement_type)
    result = query.order_by(models.Kardex.date.desc()).limit(min(payload.limit or 5000, 20000)).all()
    rows = []
    for kardex, product, warehouse in result:
        movement = _export_enum(kardex.movement_type)
        values = {
            "date": _export_date(kardex.date),
            "product": product.name if product else "",
            "sku": product.sku if product else "",
            "movement_type": MOVEMENT_LABELS_EXPORT.get(movement, movement),
            "quantity": _export_number(kardex.quantity),
            "balance_after": _export_number(kardex.balance_after),
            "warehouse": warehouse.name if warehouse else "",
            "description": kardex.description or "",
        }
        rows.append({KARDEX_EXPORT_COLUMNS[col]: values.get(col, "") for col in columns if col in KARDEX_EXPORT_COLUMNS})
    return rows, len(rows)


def _export_serial_rows(db: Session, payload: InventoryExportRequest, columns: List[str]):
    query = db.query(models.ProductInstance, models.Product, models.Warehouse).join(
        models.Product, models.ProductInstance.product_id == models.Product.id
    ).outerjoin(models.Warehouse, models.ProductInstance.warehouse_id == models.Warehouse.id)
    if payload.search:
        tokens = [token for token in payload.search.strip().split() if token]
        for token in tokens:
            like = f"%{token}%"
            query = query.filter(or_(models.Product.name.ilike(like), models.Product.sku.ilike(like), models.ProductInstance.serial_number.ilike(like), models.ProductInstance.color_name.ilike(like)))
    if payload.category_id:
        query = query.filter(models.Product.category_id == payload.category_id)
    if payload.warehouse_id:
        query = query.filter(models.ProductInstance.warehouse_id == payload.warehouse_id)
    rows_data = query.order_by(func.lower(models.Product.name), models.ProductInstance.serial_number).limit(min(payload.limit or 5000, 20000)).all()
    rows = []
    for instance, product, warehouse in rows_data:
        status_value = _export_enum(instance.status)
        values = {
            "product": product.name if product else "",
            "sku": product.sku if product else "",
            "serial_number": instance.serial_number,
            "status": status_value,
            "color_name": instance.color_name or "",
            "color_hex": instance.color_hex or "",
            "warehouse": warehouse.name if warehouse else "",
            "cost": _export_number(instance.cost),
            "created_at": _export_date(instance.created_at),
        }
        rows.append({SERIAL_EXPORT_COLUMNS[col]: values.get(col, "") for col in columns if col in SERIAL_EXPORT_COLUMNS})
    return rows, len(rows)


def _export_sheet_name(export_type: str) -> str:
    return {
        "catalog_basic": "Catalogo",
        "catalog_prices": "Catalogo precios",
        "stock": "Stock",
        "prices": "Precios",
        "kardex": "Kardex",
        "serials": "Seriales",
    }.get(export_type, "Inventario")[:31]


def _prepare_inventory_export(db: Session, payload: InventoryExportRequest, current_user: models.User):
    export_type = payload.export_type or "catalog_basic"
    required_permission = EXPORT_PERMISSION_BY_TYPE.get(export_type)
    if not required_permission:
        raise HTTPException(status_code=400, detail="Tipo de exportacion no soportado")

    from ..services.permissions_service import user_has_permission
    if not user_has_permission(db, current_user, required_permission):
        raise HTTPException(status_code=403, detail="No tienes permisos para esta descarga")

    allowed_columns = (
        KARDEX_EXPORT_COLUMNS if export_type == "kardex"
        else SERIAL_EXPORT_COLUMNS if export_type == "serials"
        else PRODUCT_EXPORT_COLUMNS
    )
    requested = [col for col in (payload.columns or DEFAULT_EXPORT_COLUMNS.get(export_type, [])) if col in allowed_columns]
    if not requested:
        requested = DEFAULT_EXPORT_COLUMNS.get(export_type, list(allowed_columns.keys()))

    include_lists = payload.include_price_lists or export_type in {"catalog_prices", "prices"}
    if export_type in {"catalog_basic", "catalog_prices", "prices"}:
        rows, total = _export_catalog_rows(db, payload, requested, include_price_lists=include_lists)
    elif export_type == "stock":
        rows, total = _export_stock_rows(db, payload, requested)
    elif export_type == "kardex":
        rows, total = _export_kardex_rows(db, payload, requested)
    elif export_type == "serials":
        rows, total = _export_serial_rows(db, payload, requested)
    else:
        raise HTTPException(status_code=400, detail="Tipo de exportacion no soportado")

    return {
        "export_type": export_type,
        "rows": rows,
        "total": total,
        "requested_columns": requested,
        "sheet_name": _export_sheet_name(export_type),
        "format": (payload.format or "xlsx").lower(),
    }


def _excel_response(rows: List[dict], export_type: str, total: int, payload: InventoryExportRequest):
    from io import BytesIO
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    sheet_name = _export_sheet_name(export_type)
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        meta = pd.DataFrame([
            {"Campo": "Tipo", "Valor": export_type},
            {"Campo": "Generado", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M")},
            {"Campo": "Registros", "Valor": total},
            {"Campo": "Busqueda", "Valor": payload.search or ""},
            {"Campo": "Fecha inicio", "Valor": payload.start_date or ""},
            {"Campo": "Fecha fin", "Valor": payload.end_date or ""},
        ])
        meta.to_excel(writer, index=False, sheet_name="Resumen")
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        for ws in workbook.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F46E5")
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for column in ws.columns:
                max_len = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    max_len = max(max_len, len(str(cell.value or "")))
                ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 42)
    output.seek(0)
    filename = f"inventario_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _csv_response(rows: List[dict], export_type: str):
    from io import BytesIO
    import pandas as pd

    output = BytesIO()
    df = pd.DataFrame(rows)
    output.write(df.to_csv(index=False).encode("utf-8-sig"))
    output.seek(0)
    filename = f"inventario_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/export/preview", dependencies=[Depends(require_any_permission([
    "inventory.products.view",
    "inventory.kardex.view",
    "inventory.serials.view",
]))])
def preview_inventory_export(
    payload: InventoryExportRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    prepared = _prepare_inventory_export(db, payload, current_user)
    sample = prepared["rows"][:5]
    return {
        "export_type": prepared["export_type"],
        "format": prepared["format"],
        "total_rows": prepared["total"],
        "columns": list(sample[0].keys()) if sample else [],
        "column_count": len(list(sample[0].keys())) if sample else len(prepared["requested_columns"]),
        "sample": sample,
        "limit": min(payload.limit or 5000, 20000),
        "limited": prepared["total"] >= min(payload.limit or 5000, 20000),
    }


@router.post("/export/modular", dependencies=[Depends(require_any_permission([
    "inventory.products.view",
    "inventory.kardex.view",
    "inventory.serials.view",
]))])
def export_inventory_modular(
    payload: InventoryExportRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    prepared = _prepare_inventory_export(db, payload, current_user)
    if prepared["format"] == "csv":
        return _csv_response(prepared["rows"], prepared["export_type"])
    return _excel_response(prepared["rows"], prepared["export_type"], prepared["total"], payload)

# --- INTER-COMPANY TRANSFER ENDPOINTS ---
from fastapi import UploadFile, File, Body
from ..services.inventory_service import InventoryService
from ..schemas import TransferPackageSchema, TransferResultSchema, TransferPreviewResult, TransferImportV2Request
class TransferRequest(BaseModel):
    items: List[Dict[str, Any]]
    source_company: str
    warehouse_id: Optional[int] = None
    destination_company: Optional[str] = None
    dispatch_notes: Optional[str] = None
    photo_urls: Optional[List[str]] = None

class DispatchGuideRequest(BaseModel):
    package: Dict[str, Any]
    source_company: Optional[str] = None
    destination_company: Optional[str] = None
    notes: Optional[str] = None

@router.post("/transfer/export", response_model=TransferPackageSchema, dependencies=[Depends(require_permission("inventory.transfers.export"))])
def export_transfer_package(
    request: TransferRequest,
    db: Session = Depends(get_db),
    # user: models.User = Depends(get_current_active_user) # Uncomment when security is ready
):
    """
    Generate a JSON package for transfer.
    Deducts stock immediately from this instance.
    """
    try:
        return InventoryService.generate_transfer_package_v2(
            db,
            request.items,
            request.source_company,
            request.warehouse_id,
            request.photo_urls,
            request.destination_company,
            request.dispatch_notes
        )
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred during transfer export: {e}")


def _business_config_map(db: Session) -> Dict[str, str]:
    try:
        rows = db.query(models.BusinessConfig).filter(
            models.BusinessConfig.key.in_([
                "business_name", "business_rif", "business_address", "business_phone"
            ])
        ).all()
        return {row.key: row.value for row in rows if row.value}
    except Exception:
        return {}


def _safe_transfer_filename(value: str) -> str:
    raw = (value or "guia-despacho").strip().lower()
    cleaned = "".join(ch if ch.isalnum() else "-" for ch in raw)
    while "--" in cleaned:
        cleaned = cleaned.replace("--", "-")
    return cleaned.strip("-") or "guia-despacho"


@router.post("/transfer/dispatch-guide", dependencies=[Depends(require_permission("inventory.transfers.export"))])
def generate_transfer_dispatch_guide(
    request: DispatchGuideRequest,
    db: Session = Depends(get_db),
):
    """Generate a printable dispatch guide from an inter-company transfer package."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    package = request.package or {}
    items = package.get("items") or []
    if not isinstance(items, list) or not items:
        raise HTTPException(status_code=400, detail="El paquete no tiene productos para generar la guia")

    cfg = _business_config_map(db)
    now = datetime.now()
    package_id = str(package.get("package_id") or "SIN-PAQUETE")
    guide_number = str(
        package.get("dispatch_guide_number")
        or f"GD-{now.strftime('%Y%m%d')}-{package_id[-8:].upper()}"
    )
    source_company = request.source_company or package.get("source_company") or cfg.get("business_name") or "Mi Inventario"
    destination_company = request.destination_company or package.get("destination_company") or "Destino por definir"
    notes = request.notes or package.get("dispatch_notes") or ""
    generated_at = package.get("generated_at") or now.isoformat()

    models_count = int(package.get("models_count") or package.get("items_count") or len(items))
    units_count = sum(float(item.get("quantity") or 0) for item in items)
    imei_count = sum(len(item.get("serial_numbers") or []) for item in items)
    photos_count = int(package.get("photos_count") or len(package.get("photo_urls") or []))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DispatchTitle", parent=styles["Title"], fontSize=16, leading=19,
        textColor=colors.HexColor("#1E3A8A"), alignment=TA_CENTER, spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "DispatchSubtitle", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.HexColor("#64748B"), alignment=TA_CENTER,
    )
    label_style = ParagraphStyle(
        "DispatchLabel", parent=styles["Normal"], fontSize=7, leading=9,
        textColor=colors.HexColor("#64748B"), fontName="Helvetica-Bold",
    )
    value_style = ParagraphStyle(
        "DispatchValue", parent=styles["Normal"], fontSize=9, leading=11,
        textColor=colors.HexColor("#0F172A"), fontName="Helvetica-Bold",
    )
    small_style = ParagraphStyle(
        "DispatchSmall", parent=styles["Normal"], fontSize=7, leading=9,
        textColor=colors.HexColor("#475569"),
    )
    right_style = ParagraphStyle(
        "DispatchRight", parent=value_style, alignment=TA_RIGHT,
    )

    def cell_label(text):
        return Paragraph(str(text or ""), label_style)

    def cell_value(text):
        return Paragraph(str(text or ""), value_style)

    story = []
    story.append(Paragraph("GUIA DE DESPACHO", title_style))
    story.append(Paragraph("Documento de control logistico. No sustituye factura fiscal.", subtitle_style))
    story.append(Spacer(1, 7 * mm))

    header = Table([
        [cell_label("EMISOR"), cell_label("GUIA"), cell_label("FECHA")],
        [cell_value(source_company), Paragraph(guide_number, right_style), Paragraph(now.strftime("%d/%m/%Y %H:%M"), right_style)],
    ], colWidths=[95 * mm, 45 * mm, 40 * mm])
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#475569")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(header)
    story.append(Spacer(1, 5 * mm))

    info = Table([
        [cell_label("DESTINO / EMPRESA RECEPTORA"), cell_label("ALMACEN ORIGEN"), cell_label("PAQUETE")],
        [cell_value(destination_company), cell_value(package.get("source_warehouse_name") or "Sin almacen"), cell_value(package_id)],
        [cell_label("GENERADO EN"), cell_label("CONTACTO / RIF"), cell_label("NOTAS")],
        [cell_value(generated_at), cell_value(cfg.get("business_rif") or cfg.get("business_phone") or ""), Paragraph(notes or "Sin notas", small_style)],
    ], colWidths=[65 * mm, 55 * mm, 60 * mm])
    info.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F8FAFC")),
        ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(info)
    story.append(Spacer(1, 5 * mm))

    totals = Table([[f"Modelos: {models_count}", f"Unidades: {units_count:g}", f"Seriales/IMEI: {imei_count}", f"Fotos: {photos_count}"]], colWidths=[45 * mm] * 4)
    totals.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ECFDF5")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#047857")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#A7F3D0")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(totals)
    story.append(Spacer(1, 5 * mm))

    table_data = [[
        Paragraph("SKU", label_style),
        Paragraph("PRODUCTO", label_style),
        Paragraph("CANT.", label_style),
        Paragraph("SERIALES / IMEI", label_style),
    ]]
    for item in items:
        serials = item.get("serial_numbers") or []
        serial_text = ", ".join(str(s) for s in serials) if serials else "-"
        table_data.append([
            Paragraph(str(item.get("sku") or ""), small_style),
            Paragraph(str(item.get("name") or ""), small_style),
            Paragraph(f"{float(item.get('quantity') or 0):g}", small_style),
            Paragraph(serial_text, small_style),
        ])

    details = Table(table_data, colWidths=[32 * mm, 73 * mm, 20 * mm, 55 * mm], repeatRows=1)
    details.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (2, 1), (2, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(details)
    story.append(Spacer(1, 10 * mm))

    signatures = Table([
        ["ENTREGA", "TRANSPORTA", "RECIBE"],
        ["\n\nFirma / Sello", "\n\nNombre / Cedula", "\n\nFirma / Sello"],
    ], colWidths=[60 * mm] * 3)
    signatures.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 1), (-1, 1), 18),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
    ]))
    story.append(signatures)

    doc.build(story)
    buffer.seek(0)
    filename = _safe_transfer_filename(f"guia-despacho-{guide_number}") + ".pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.post("/transfer/import", response_model=TransferResultSchema, dependencies=[Depends(require_permission("inventory.transfers.import"))])
async def import_transfer_package(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Import a JSON transfer package (legacy - exact SKU match only).
    Adds stock to this instance.
    """
    content = await file.read()
    return InventoryService.process_transfer_package(db, content)

@router.post("/transfer/preview", response_model=TransferPreviewResult, dependencies=[Depends(require_permission("inventory.transfers.import"))])
async def preview_transfer(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Preview a JSON transfer package before importing.
    Returns match results for each item using a 4-step cascade:
    exact SKU, case-insensitive SKU, normalized SKU, name ILIKE.
    """
    content = await file.read()
    return InventoryService.preview_transfer_package(db, content)

@router.post("/transfer/import-mapped", dependencies=[Depends(require_permission("inventory.transfers.import"))])
async def import_transfer_mapped(
    request: TransferImportV2Request = Body(...),
    db: Session = Depends(get_db)
):
    """
    Import a transfer package with explicit product mappings (v2).
    Each item can be mapped to an existing product_id or flagged to create a new product.
    """
    data = request.model_dump()
    return InventoryService.process_transfer_package_v2(db, data, request.warehouse_id)

@router.post("/transfer/upload-photo", dependencies=[Depends(require_any_permission(["inventory.transfers.export", "inventory.transfers.import"]))])
async def upload_transfer_photo(file: UploadFile = File(...)):
    """
    Upload a photo as evidence for a transfer package.
    Photos are stored permanently in /app/media/transfers/.
    Returns the public URL of the uploaded image.
    """
    from ..utils.media_utils import save_upload_file
    url = save_upload_file(file, folder="transfers")
    return {"url": url}

@router.post("/bulk-entry", dependencies=[Depends(warehouse_or_admin)])
def bulk_entry(
    entry_data: schemas.SerializedEntry, 
    db: Session = Depends(get_db)
):
    """
    Mass entry of serialized items (IMEIs).
    Optimized for performance ("Metralleta").
    """
    return InventoryService.process_bulk_entry(db, entry_data)

@router.get("/validate-imei")
def validate_imei(product_id: int, imei: str, db: Session = Depends(get_db)):
    """
    Check if an IMEI is valid and available for a given product.
    """
    return InventoryService.validate_imei_availability(db, product_id, imei)

@router.get("/validate-entry")
def validate_imei_for_entry(imei: str, db: Session = Depends(get_db)):
    """
    Check if an IMEI is ALREADY in the database.
    Used for Reception (Entry) to prevent duplicates.
    Returns: {"exists": bool, "message": str}
    """
    return InventoryService.validate_imei_for_entry(db, imei)


@router.get("/serialized-instances", dependencies=[Depends(require_permission("inventory.serials.view"))])
def get_all_serialized_instances(db: Session = Depends(get_db)):
    """
    Get ALL serialized instances (IMEIs) across all products.
    Used for the serialized report PDF.
    """
    instances = db.query(models.ProductInstance).options(
        joinedload(models.ProductInstance.warehouse),
        joinedload(models.ProductInstance.product)
    ).order_by(
        models.ProductInstance.product_id,
        models.ProductInstance.status,
        models.ProductInstance.created_at.desc()
    ).all()
    return instances

@router.get("/product/{product_id}/instances", dependencies=[Depends(require_permission("inventory.serials.view"))])
def get_product_instances(product_id: int, db: Session = Depends(get_db)):
    """
    Get all serialized instances (IMEIs) for a specific product.
    Includes status, warehouse, and dates.
    """
    instances = db.query(models.ProductInstance).options(
        joinedload(models.ProductInstance.warehouse)
    ).filter(
        models.ProductInstance.product_id == product_id
    ).order_by(models.ProductInstance.status, models.ProductInstance.created_at.desc()).all()

    return instances


@router.delete("/instance/{instance_id}", dependencies=[Depends(require_permission("inventory.serials.delete"))])
def delete_imei_instance(
    instance_id: int,
    reason: str = "Corrección de error",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_permission("inventory.serials.delete"))
):
    """
    Eliminar un IMEI/serial ingresado por error.
    Solo se permite borrar físicamente seriales disponibles y sin historial.
    Si el IMEI ya participó en ventas/devoluciones, se conserva la trazabilidad.
    """
    instance = db.query(models.ProductInstance).filter(
        models.ProductInstance.id == instance_id
    ).options(
        joinedload(models.ProductInstance.product),
        joinedload(models.ProductInstance.warehouse)
    ).first()

    if not instance:
        raise HTTPException(status_code=404, detail="IMEI no encontrado")

    product = instance.product
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    result = {
        "imei": instance.serial_number,
        "product_name": product.name,
        "status_was": instance.status.value if hasattr(instance.status, 'value') else str(instance.status),
        "stock_adjusted": False
    }

    sale_link = db.query(models.SaleDetailInstance, models.SaleDetail, models.Sale).join(
        models.SaleDetail, models.SaleDetail.id == models.SaleDetailInstance.sale_detail_id
    ).join(
        models.Sale, models.Sale.id == models.SaleDetail.sale_id
    ).filter(
        models.SaleDetailInstance.product_instance_id == instance.id
    ).first()

    return_link = db.query(models.ReturnDetailInstance, models.ReturnDetail, models.Return).join(
        models.ReturnDetail, models.ReturnDetail.id == models.ReturnDetailInstance.return_detail_id
    ).join(
        models.Return, models.Return.id == models.ReturnDetail.return_id
    ).filter(
        models.ReturnDetailInstance.product_instance_id == instance.id
    ).first()

    if sale_link or return_link:
        sale_id = sale_link[2].id if sale_link else None
        return_id = return_link[2].id if return_link else None
        references = []
        if sale_id:
            references.append(f"venta #{sale_id}")
        if return_id:
            references.append(f"devolución #{return_id}")
        raise HTTPException(
            status_code=409,
            detail={
                "message": (
                    f"No se puede eliminar el IMEI {instance.serial_number} porque ya tiene historial en "
                    f"{', '.join(references)}. Para corregirlo, usa devolución/anulación o ajuste de estado, "
                    "pero no borrado físico."
                ),
                "imei": instance.serial_number,
                "product_name": product.name,
                "status": result["status_was"],
                "sale_id": sale_id,
                "return_id": return_id,
                "code": "IMEI_HAS_HISTORY"
            }
        )

    if str(instance.status) not in ("AVAILABLE", "ProductInstanceStatus.AVAILABLE"):
        raise HTTPException(
            status_code=409,
            detail={
                "message": f"Solo se pueden eliminar IMEIs disponibles y sin historial. El IMEI {instance.serial_number} está en estado {result['status_was']}.",
                "imei": instance.serial_number,
                "product_name": product.name,
                "status": result["status_was"],
                "code": "IMEI_NOT_AVAILABLE"
            }
        )

    if str(instance.status) in ("AVAILABLE", "ProductInstanceStatus.AVAILABLE"):
        # Descontar del stock ya que el IMEI estaba disponible
        product.stock = max(0, float(product.stock) - 1)

        # Descontar también del product_stocks del almacén
        if instance.warehouse_id:
            ps = db.query(models.ProductStock).filter(
                models.ProductStock.product_id == product.id,
                models.ProductStock.warehouse_id == instance.warehouse_id
            ).first()
            if ps:
                ps.quantity = max(0, float(ps.quantity) - 1)

        # Kardex de ajuste
        kardex = models.Kardex(
            product_id=product.id,
            warehouse_id=instance.warehouse_id,
            movement_type="ADJUSTMENT_OUT",
            quantity=-1,
            balance_after=product.stock,
            description=f"Eliminación IMEI {instance.serial_number} — {reason}",
            date=datetime.now()
        )
        db.add(kardex)
        result["stock_adjusted"] = True

    # Eliminar el instance
    db.delete(instance)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail={
                "message": f"No se puede eliminar el IMEI {instance.serial_number} porque está relacionado con historial del sistema.",
                "imei": instance.serial_number,
                "product_name": product.name,
                "status": result["status_was"],
                "code": "IMEI_DELETE_CONFLICT"
            }
        )

    return {"status": "deleted", **result}


@router.patch("/instance/{instance_id}/fix-serial", dependencies=[Depends(warehouse_or_admin)])
def fix_imei_serial(
    instance_id: int,
    body: dict,
    db: Session = Depends(get_db)
):
    """Corregir el número de serial/IMEI de un registro existente"""
    instance = db.query(models.ProductInstance).filter(
        models.ProductInstance.id == instance_id
    ).first()
    if not instance:
        raise HTTPException(status_code=404, detail="IMEI no encontrado")

    new_serial = body.get("serial_number", "").strip()
    if not new_serial:
        raise HTTPException(status_code=400, detail="Serial no puede estar vacío")

    # Verificar que no exista otro con ese serial
    existing = db.query(models.ProductInstance).filter(
        models.ProductInstance.serial_number == new_serial,
        models.ProductInstance.id != instance_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"El serial {new_serial} ya existe en el sistema")

    old_serial = instance.serial_number
    instance.serial_number = new_serial
    db.commit()

    return {
        "status": "updated",
        "old_serial": old_serial,
        "new_serial": new_serial,
        "product_id": instance.product_id
    }


@router.get("/lookup-imei", dependencies=[Depends(require_any_permission(["inventory.serials.view", "pos.sell"]))])
def lookup_imei(imei: str, db: Session = Depends(get_db)):
    """
    Buscar un producto por IMEI/serial.
    Devuelve el producto, almacén y estado del IMEI.
    """
    instance = db.query(models.ProductInstance).filter(
        models.ProductInstance.serial_number.ilike(imei.strip())
    ).options(
        joinedload(models.ProductInstance.product),
        joinedload(models.ProductInstance.warehouse),
    ).first()

    if not instance:
        raise HTTPException(status_code=404, detail="IMEI no encontrado en el inventario")

    product = instance.product
    available_stock = 0
    if product and getattr(product, "has_imei", False):
        available_stock = db.query(func.count(models.ProductInstance.id)).filter(
            models.ProductInstance.product_id == product.id,
            models.ProductInstance.status == models.ProductInstanceStatus.AVAILABLE,
        ).scalar() or 0
    elif product:
        available_stock = product.stock or 0

    return {
        "imei": instance.serial_number,
        "instance_id": instance.id,
        "status": instance.status,
        "warehouse": instance.warehouse.name if instance.warehouse else None,
        "color_name": getattr(instance, "color_name", None),
        "color_hex": getattr(instance, "color_hex", None),
        "product": {
            "id": product.id,
            "name": product.name,
            "sku": product.sku,
            "price": float(product.price),
            "cost": float(getattr(product, "cost_price", None) or getattr(product, "cost", None) or 0),
            "stock": float(available_stock),
            "category_name": product.category.name if product.category else None,
            "image_url": product.image_url,
            "description": product.description,
            "has_imei": product.has_imei,
        } if product else None,
        "sold_at": instance.updated_at.isoformat() if instance.status == "SOLD" and instance.updated_at else None,
    }
