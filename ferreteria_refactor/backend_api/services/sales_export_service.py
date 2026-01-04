
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
import pandas as pd
import io
from datetime import datetime, date
from decimal import Decimal
from ..models import models
from ..utils.payment_utils import normalize_payment_method, get_currency_symbol, normalize_currency_code

def generate_sales_excel(db: Session, start_date: date, end_date: date) -> io.BytesIO:
    """
    Generates a professional Excel file with:
    - Tab 1: "Resumen Z" (Z-Report Summary) with cash reconciliation, payment breakdown, category sales
    - Tab 2: "Ventas Detalladas" (Detailed Sales) with explicit currency columns
    
    Enhanced with professional formatting: bold headers, colors, borders, frozen panes
    """
    
    # Date boundaries
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # ============================================
    # QUERY DATA
    # ============================================
    
    # 1. Fetch Sales
    sales = db.query(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).order_by(models.Sale.date.desc()).all()
    
    # 2. Fetch Payment Breakdown
    payments_raw = db.query(
        models.SalePayment.payment_method,
        models.SalePayment.currency,
        func.sum(models.SalePayment.amount).label('total'),
        func.count(models.SalePayment.id).label('count')
    ).join(models.Sale).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).group_by(
        models.SalePayment.payment_method,
        models.SalePayment.currency
    ).all()
    
    # 3. Fetch Category Sales
    category_sales = db.query(
        models.Category.name.label('category_name'),
        func.sum(models.SaleDetail.subtotal).label('total_usd')
    ).join(
        models.Product, models.SaleDetail.product_id == models.Product.id
    ).join(
        models.Category, models.Product.category_id == models.Category.id
    ).join(
        models.Sale, models.SaleDetail.sale_id == models.Sale.id
    ).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).group_by(
        models.Category.name
    ).all()
    
    # 4. Find Cash Session
    cash_session = db.query(models.CashSession).filter(
        models.CashSession.start_time >= start_dt,
        models.CashSession.start_time <= end_dt
    ).first()
    
    # ============================================
    # TAB 1: RESUMEN Z (Z-REPORT SUMMARY)
    # ============================================
    
    # Section 1: Cash Reconciliation
    cash_usd_sales = Decimal("0.00")
    cash_ves_sales = Decimal("0.00")
    
    for p in payments_raw:
        method = normalize_payment_method(p[0])
        currency = p[1] or "USD"
        amount = Decimal(str(p[2]))
        
        if method == "Efectivo":
            if get_currency_symbol(currency) == "Bs":
                cash_ves_sales += amount
            else:
                cash_usd_sales += amount
    
    # Build cash reconciliation data
    if cash_session:
        movements = db.query(models.CashMovement).filter(
            models.CashMovement.session_id == cash_session.id
        ).all()
        
        deposits_usd = sum((Decimal(str(m.amount)) for m in movements 
                           if m.type == "DEPOSIT" and m.currency == "USD"), Decimal("0.00"))
        expenses_usd = sum((Decimal(str(m.amount)) for m in movements 
                           if m.type in ["EXPENSE", "WITHDRAWAL", "OUT"] and m.currency == "USD"), Decimal("0.00"))
        deposits_ves = sum((Decimal(str(m.amount)) for m in movements 
                           if m.type == "DEPOSIT" and m.currency in ["VES", "Bs", "VEF"]), Decimal("0.00"))
        expenses_ves = sum((Decimal(str(m.amount)) for m in movements 
                           if m.type in ["EXPENSE", "WITHDRAWAL", "OUT"] and m.currency in ["VES", "Bs", "VEF"]), Decimal("0.00"))
        
        initial_usd = Decimal(str(cash_session.initial_cash or 0))
        initial_ves = Decimal(str(cash_session.initial_cash_bs or 0))
        
        cash_recon_data = {
            'Concepto': ['Fondo Inicial', 'Ventas en Efectivo', 'Depósitos', 'Retiros/Gastos', 'ESPERADO EN GAVETA'],
            'USD': [
                f"${float(initial_usd):,.2f}",
                f"${float(cash_usd_sales):,.2f}",
                f"${float(deposits_usd):,.2f}",
                f"${float(expenses_usd):,.2f}",
                f"${float(initial_usd + cash_usd_sales + deposits_usd - expenses_usd):,.2f}"
            ],
            'Bs': [
                f"Bs {float(initial_ves):,.2f}",
                f"Bs {float(cash_ves_sales):,.2f}",
                f"Bs {float(deposits_ves):,.2f}",
                f"Bs {float(expenses_ves):,.2f}",
                f"Bs {float(initial_ves + cash_ves_sales + deposits_ves - expenses_ves):,.2f}"
            ]
        }
    else:
        # No cash session - simplified view
        cash_recon_data = {
            'Concepto': ['Ventas en Efectivo', 'ESPERADO EN GAVETA'],
            'USD': [
                f"${float(cash_usd_sales):,.2f}",
                f"${float(cash_usd_sales):,.2f}"
            ],
            'Bs': [
                f"Bs {float(cash_ves_sales):,.2f}",
                f"Bs {float(cash_ves_sales):,.2f}"
            ]
        }
    
    df_cash_recon = pd.DataFrame(cash_recon_data)
    
    # Section 2: Payment Method Breakdown
    payment_breakdown_data = []
    for p in payments_raw:
        method = normalize_payment_method(p[0])
        currency = p[1] or "USD"
        symbol = get_currency_symbol(currency)
        amount = float(p[2])
        count = p[3]
        
        payment_breakdown_data.append({
            'Método': method,
            'Moneda': currency if symbol == "$" else "VES",
            'Monto': f"{symbol} {amount:,.2f}",
            '# Transacciones': count
        })
    
    df_payment_breakdown = pd.DataFrame(payment_breakdown_data) if payment_breakdown_data else pd.DataFrame()
    
    # Section 3: Category Sales
    category_data = []
    for cat in category_sales:
        category_data.append({
            'Categoría': cat.category_name or 'Sin Categoría',
            'Monto USD': f"${float(cat.total_usd or 0):,.2f}"
        })
    
    df_category = pd.DataFrame(category_data) if category_data else pd.DataFrame()
    
    # ============================================
    # TAB 2: VENTAS DETALLADAS (DETAILED SALES)
    # ============================================
    
    detailed_sales_data = []
    for sale in sales:
        if hasattr(sale, 'status') and sale.status == "VOIDED":
            continue
        
        ticket_id = f"#{sale.id:06d}"
        customer_name = sale.customer.name if sale.customer else "Cliente General"
        total_usd = float(sale.total_amount)
        
        # Calculate cost & profit
        total_cost = sum((float(d.cost_at_sale or 0) * float(d.quantity)) for d in sale.details)
        real_profit = total_usd - total_cost
        
        # Currency info
        total_ves = float(sale.total_amount_bs or 0)
        implied_rate = (total_ves / total_usd) if total_usd > 0 and total_ves > 0 else 0.0
        
        # Change
        change_amt = float(sale.change_amount or 0)
        change_curr = sale.change_currency or "VES"
        change_str = f"{change_curr} {change_amt:,.2f}" if change_amt > 0 else "N/A"
        
        # Payment methods
        payments_list = []
        for p in sale.payments:
            curr_sym = "$" if p.currency in ["USD", "$"] else "Bs"
            p_amt = float(p.amount)
            payments_list.append(f"{normalize_payment_method(p.payment_method)}: {curr_sym}{p_amt:,.2f}")
        payment_methods_str = " | ".join(payments_list)
        
        detailed_sales_data.append({
            "Fecha/Hora": sale.date,
            "# Ticket": ticket_id,
            "Cliente": customer_name,
            "Total Venta (USD)": f"${total_usd:,.2f}",
            "Costo Total (USD)": f"${total_cost:,.2f}",
            "Ganancia Real (USD)": f"${real_profit:,.2f}",
            "Total Cobrado (Bs)": f"Bs {total_ves:,.2f}",
            "Tasa Implícita": f"{implied_rate:,.2f}",
            "Métodos de Pago": payment_methods_str,
            "Vuelto Entregado": change_str
        })
    
    df_detailed = pd.DataFrame(detailed_sales_data) if detailed_sales_data else pd.DataFrame(columns=[
        "Fecha/Hora", "# Ticket", "Cliente", "Total Venta (USD)", "Costo Total (USD)", 
        "Ganancia Real (USD)", "Total Cobrado (Bs)", "Tasa Implícita", "Métodos de Pago", "Vuelto Entregado"
    ])
    
    # ============================================
    # CREATE EXCEL FILE
    # ============================================
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # TAB 1: Resumen Z
        start_row = 0
        
        # Write Cash Reconciliation
        df_cash_recon.to_excel(writer, sheet_name='Resumen Z', index=False, startrow=start_row)
        start_row += len(df_cash_recon) + 3
        
        # Write Payment Breakdown
        if not df_payment_breakdown.empty:
            worksheet = writer.sheets['Resumen Z']
            worksheet.cell(row=start_row + 1, column=1, value='VENTAS POR MÉTODO DE PAGO')
            df_payment_breakdown.to_excel(writer, sheet_name='Resumen Z', index=False, startrow=start_row + 1)
            start_row += len(df_payment_breakdown) + 4
        
        # Write Category Sales
        if not df_category.empty:
            worksheet = writer.sheets['Resumen Z']
            worksheet.cell(row=start_row + 1, column=1, value='VENTAS POR CATEGORÍA')
            df_category.to_excel(writer, sheet_name='Resumen Z', index=False, startrow=start_row + 1)
        
        # TAB 2: Ventas Detalladas
        df_detailed.to_excel(writer, sheet_name='Ventas Detalladas', index=False)
        
        # ============================================
        # APPLY PROFESSIONAL FORMATTING
        # ============================================
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=12)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format Resumen Z sheet
        ws_resumen = writer.sheets['Resumen Z']
        
        # Freeze top row
        ws_resumen.freeze_panes = 'A2'
        
        # Format headers for cash reconciliation (row 1)
        for col in range(1, len(df_cash_recon.columns) + 1):
            cell = ws_resumen.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        
        # Format data rows for cash reconciliation
        for row in range(2, len(df_cash_recon) + 2):
            for col in range(1, len(df_cash_recon.columns) + 1):
                cell = ws_resumen.cell(row=row, column=col)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
                # Bold the last row (ESPERADO EN GAVETA)
                if row == len(df_cash_recon) + 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Format section headers
        payment_header_row = len(df_cash_recon) + 4
        if not df_payment_breakdown.empty:
            # Section title
            cell = ws_resumen.cell(row=payment_header_row, column=1)
            cell.fill = section_fill
            cell.font = section_font
            
            # Column headers
            for col in range(1, len(df_payment_breakdown.columns) + 1):
                cell = ws_resumen.cell(row=payment_header_row + 1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
            
            # Data rows
            for row in range(payment_header_row + 2, payment_header_row + len(df_payment_breakdown) + 2):
                for col in range(1, len(df_payment_breakdown.columns) + 1):
                    cell = ws_resumen.cell(row=row, column=col)
                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
        
        # Format category section
        category_header_row = payment_header_row + len(df_payment_breakdown) + 5 if not df_payment_breakdown.empty else len(df_cash_recon) + 4
        if not df_category.empty:
            # Section title
            cell = ws_resumen.cell(row=category_header_row, column=1)
            cell.fill = section_fill
            cell.font = section_font
            
            # Column headers
            for col in range(1, len(df_category.columns) + 1):
                cell = ws_resumen.cell(row=category_header_row + 1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
            
            # Data rows
            for row in range(category_header_row + 2, category_header_row + len(df_category) + 2):
                for col in range(1, len(df_category.columns) + 1):
                    cell = ws_resumen.cell(row=row, column=col)
                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
        
        # Format Ventas Detalladas sheet
        ws_detailed = writer.sheets['Ventas Detalladas']
        
        # Freeze top row
        ws_detailed.freeze_panes = 'A2'
        
        # Format headers
        for col in range(1, len(df_detailed.columns) + 1):
            cell = ws_detailed.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_thin
        
        # Format data rows with alternating colors
        for row in range(2, len(df_detailed) + 2):
            row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if row % 2 == 0 else None
            for col in range(1, len(df_detailed.columns) + 1):
                cell = ws_detailed.cell(row=row, column=col)
                cell.border = border_thin
                if row_fill:
                    cell.fill = row_fill
                # Right-align numeric columns
                if col >= 4:  # Numeric columns start at column 4
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths for both sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def generate_product_sales_excel(db: Session, start_date: date, end_date: date) -> io.BytesIO:
    """
    Generates an Excel file for Product Sales:
    - Columns: Product Name, Category, Quantity Sold, Total Revenue, Avg Price
    """
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # Query Data
    query = db.query(
        models.Product.name.label('product_name'),
        models.Category.name.label('category_name'),
        func.sum(models.SaleDetail.quantity).label('total_qty'),
        func.sum(models.SaleDetail.subtotal).label('total_rev')
    ).join(
        models.SaleDetail, models.Product.id == models.SaleDetail.product_id
    ).join(
        models.Sale, models.SaleDetail.sale_id == models.Sale.id
    ).outerjoin(
        models.Category, models.Product.category_id == models.Category.id
    ).filter(
        models.Sale.date >= start_dt,
        models.Sale.date <= end_dt
    ).group_by(
        models.Product.id,
        models.Product.name,
        models.Category.name
    ).order_by(desc('total_rev')).all()
    
    # Build Dataframe
    data = []
    for row in query:
        qty = float(row.total_qty or 0)
        rev = float(row.total_rev or 0)
        avg = rev / qty if qty > 0 else 0
        
        data.append({
            'Producto': row.product_name,
            'Categoría': row.category_name or 'Sin Categoría',
            'Cantidad Vendida': qty,
            'Total Ventas ($)': rev,
            'Precio Promedio': avg
        })
        
    df = pd.DataFrame(data)
    
    # Create Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df.empty:
            df.to_excel(writer, sheet_name='Ventas por Producto', index=False)
            
            # Formatting
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            ws = writer.sheets['Ventas por Producto']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border_thin
                    if col >= 3: # Numeric columns
                        cell.alignment = Alignment(horizontal='right')
                        
            # Auto-width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        else:
             pd.DataFrame({'Mensaje': ['No hay ventas de productos en este período']}).to_excel(writer, sheet_name='Ventas por Producto', index=False)

    output.seek(0)
    return output
